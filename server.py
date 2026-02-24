#!/usr/bin/env python3
"""WaWi Chatbot Admin Monitor — zero-dependency server on port 7779.

Receives events from the chatbot SDK, proxies chat requests to Ollama,
manages users (block/allow), serves the admin dashboard, and pushes
live updates via SSE.

Features: cookie-based auth, admin/user roles, bot kill switch, user priorities.
"""

import ast
import base64
import json
import hashlib
import math
import os
import re
import secrets
import struct
import sys
import time
import uuid
import io
import threading
import queue
from collections import defaultdict, deque
from datetime import datetime, timezone
from http.server import BaseHTTPRequestHandler, HTTPServer, SimpleHTTPRequestHandler
from socketserver import ThreadingMixIn
from urllib.parse import urlparse, parse_qs, urlencode
from urllib.request import Request, urlopen
from urllib.error import URLError
from pathlib import Path
import sqlite3  # always available

try:
    import pyodbc
    HAS_PYODBC = True
except ImportError:
    HAS_PYODBC = False

try:
    import numpy as np
    HAS_NUMPY = True
except ImportError:
    HAS_NUMPY = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import tiktoken
    HAS_TIKTOKEN = True
except ImportError:
    HAS_TIKTOKEN = False

# ---------------------------------------------------------------------------
# OpenAI OAuth PKCE constants (from Codex CLI)
# ---------------------------------------------------------------------------
OPENAI_CLIENT_ID = "app_EMoamEEZ73f0CkXaXp7hrann"
OPENAI_AUTH_URL = "https://auth.openai.com/oauth/authorize"
OPENAI_TOKEN_URL = "https://auth.openai.com/oauth/token"
OPENAI_SCOPE = "openid profile email offline_access"
OPENAI_CALLBACK_PORT = 1455  # Must match Codex CLI's registered redirect URI

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
PORT = 7779
MAX_EVENTS = 5000
DATA_DIR = Path(__file__).resolve().parent / "data"
EVENTS_FILE = DATA_DIR / "events.jsonl"
CONFIG_FILE = DATA_DIR / "config.json"
USERS_FILE = DATA_DIR / "users.json"
ACCOUNTS_FILE = DATA_DIR / "accounts.json"
PUBLIC_DIR = Path(__file__).resolve().parent / "public"

config = {
    "provider": "ollama",  # "ollama" | "openai" | "anthropic" | "openrouter"
    "ollama_url": "http://localhost:11434",
    "default_model": "llama3:8b",
    "system_prompt": "",
    "temperature": 0.7,
    "bot_enabled": True,
    "openai_api_key": "",
    "openai_base_url": "https://api.openai.com/v1",
    "anthropic_api_key": "",
    "openrouter_api_key": "",
    "openai_oauth_token": "",
    "openai_refresh_token": "",
    "openai_token_expires": 0,
    "db_type": "",              # "" | "mssql" | "sqlite"
    "db_mssql_server": "",
    "db_mssql_database": "",
    "db_mssql_user": "",
    "db_mssql_password": "",
    "db_mssql_driver": "",
    "db_sqlite_path": "",
    "db_enabled": False,
    "db_schema_cache": "",
    "rag_enabled": False,
    "rag_top_k": 5,
    "rag_chunk_size": 500,
    "embedding_provider": "auto",  # "auto"|"openai"|"ollama"
    "embedding_model": "",
    "agent_enabled": False,
    "agent_max_steps": 8,
    "agent_web_search": True,
}
config_lock = threading.Lock()

# ---------------------------------------------------------------------------
# Global state
# ---------------------------------------------------------------------------
events: deque[dict] = deque(maxlen=MAX_EVENTS)  # O(1) append + auto-eviction
events_lock = threading.Lock()
# Pre-indexed event lookups (kept in sync with events deque)
_events_by_user: dict[str, list[dict]] = defaultdict(list)  # user_id -> [evt, ...]
_events_by_type: dict[str, int] = defaultdict(int)  # type -> count
moderation: dict[str, dict] = {}
moderation_lock = threading.Lock()
sse_clients: list[tuple[queue.Queue, dict]] = []  # (queue, session_info)
sse_lock = threading.Lock()

# User management: user_id -> { status, first_seen, last_seen, ... }
# status: "active" | "blocked" | "restricted"
users: dict[str, dict] = {}
users_lock = threading.Lock()

# Account management: username -> { username, password_hash, salt, role, priority, created_at }
accounts: dict[str, dict] = {}
accounts_lock = threading.Lock()

# Session management: session_id -> { username, role, created_at, expires_at }
sessions: dict[str, dict] = {}
sessions_lock = threading.Lock()
SESSION_TTL = 86400  # 24 hours

STOP_WORDS = frozenset(
    "der die das ein eine einer eines einem einen und oder aber wenn dann "
    "als auch noch schon doch nur mal so da hier dort wie was wer wo wann "
    "warum ist sind war waren hat hatte haben wird werden kann können muss "
    "müssen soll sollen darf dürfen möchte ich du er sie es wir ihr man "
    "mich mir dich dir sich uns euch ihm ihn ihr ihnen mein dein sein "
    "unser euer nicht kein keine mehr viel sehr gut auf aus bei mit von zu "
    "für über unter nach vor zwischen durch gegen ohne um an in im am zum "
    "zur the a an is are was were has have had will would can could shall "
    "should may might must do does did not no and or but if then than that "
    "this these those it its he she they we you i me my his her our your "
    "their what which who whom how when where why all some any much many "
    "more most other another such each every both few several than too very "
    "just also still already yet again once never always often sometimes "
    "about after before between from into through during with without of "
    "to for on at by up down in out off over under above below".split()
)


def extract_topics(text: str, top_n: int = 5) -> list[str]:
    words = re.findall(r"\b[a-zA-ZäöüÄÖÜß]{3,}\b", text.lower())
    freq: dict[str, int] = defaultdict(int)
    for w in words:
        if w not in STOP_WORDS:
            freq[w] += 1
    return [w for w, _ in sorted(freq.items(), key=lambda x: -x[1])[:top_n]]


# ---------------------------------------------------------------------------
# Shared multipart parser
# ---------------------------------------------------------------------------
def _parse_multipart(content_type: str, body: bytes) -> tuple[str, bytes]:
    """Parse multipart/form-data and extract the first file. Returns (filename, file_bytes)."""
    boundary = None
    for part in content_type.split(";"):
        part = part.strip()
        if part.startswith("boundary="):
            boundary = part[9:].strip('"')
            break
    if not boundary:
        return "", b""

    parts = body.split(("--" + boundary).encode())
    for part in parts:
        if b"Content-Disposition" not in part:
            continue
        header_end = part.find(b"\r\n\r\n")
        if header_end == -1:
            continue
        headers_raw = part[:header_end].decode("utf-8", errors="replace")
        body_raw = part[header_end + 4:]
        if body_raw.endswith(b"\r\n"):
            body_raw = body_raw[:-2]
        fn_match = re.search(r'filename="([^"]*)"', headers_raw)
        if fn_match and fn_match.group(1):
            return fn_match.group(1), body_raw
    return "", b""


# ---------------------------------------------------------------------------
# Cached tiktoken encoding
# ---------------------------------------------------------------------------
_tiktoken_encoding = None
_tiktoken_init = False


def _count_tokens(text: str) -> int:
    """Count tokens using tiktoken (cached) or fallback to char/4."""
    global _tiktoken_encoding, _tiktoken_init
    if not _tiktoken_init:
        _tiktoken_init = True
        if HAS_TIKTOKEN:
            try:
                _tiktoken_encoding = tiktoken.get_encoding("cl100k_base")
            except Exception:
                pass
    if _tiktoken_encoding is not None:
        return len(_tiktoken_encoding.encode(text))
    return len(text) // 4


# ---------------------------------------------------------------------------
# Password hashing (PBKDF2-SHA256, 260k iterations)
# ---------------------------------------------------------------------------
_PBKDF2_ITERATIONS = 260_000
_PBKDF2_PREFIX = "pbkdf2:"


def _hash_password(password: str, salt: str) -> str:
    """Hash password with PBKDF2-SHA256. Returns prefixed hash for migration detection."""
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), _PBKDF2_ITERATIONS)
    return _PBKDF2_PREFIX + dk.hex()


def _hash_password_legacy(password: str, salt: str) -> str:
    """Legacy SHA-256 hash for backward compatibility checking."""
    return hashlib.sha256((salt + password).encode("utf-8")).hexdigest()


def _verify_password(password: str, password_hash: str, salt: str) -> bool:
    """Verify password against hash. Supports both PBKDF2 and legacy SHA-256."""
    if password_hash.startswith(_PBKDF2_PREFIX):
        return secrets.compare_digest(_hash_password(password, salt), password_hash)
    # Legacy SHA-256 fallback
    return secrets.compare_digest(_hash_password_legacy(password, salt), password_hash)


def _needs_rehash(password_hash: str) -> bool:
    """Check if a stored hash needs upgrading to PBKDF2."""
    return not password_hash.startswith(_PBKDF2_PREFIX)


# ---------------------------------------------------------------------------
# Account persistence
# ---------------------------------------------------------------------------
def load_accounts():
    global accounts
    if ACCOUNTS_FILE.exists():
        try:
            with open(ACCOUNTS_FILE, "r", encoding="utf-8") as f:
                accounts = json.load(f)
        except Exception:
            pass


def save_accounts():
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with accounts_lock:
            with open(ACCOUNTS_FILE, "w", encoding="utf-8") as f:
                json.dump(accounts, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[accounts] Error saving: {e}", file=sys.stderr)


def ensure_default_admin():
    with accounts_lock:
        has_admin = any(a.get("role") == "admin" for a in accounts.values())
        if not has_admin:
            salt = secrets.token_hex(16)
            accounts["admin"] = {
                "username": "admin",
                "password_hash": _hash_password("admin", salt),
                "salt": salt,
                "role": "admin",
                "priority": "normal",
                "created_at": time.time(),
            }
    if not has_admin:
        save_accounts()
        print("[chatbot-monitor] Created default admin account (admin/admin)")


# ---------------------------------------------------------------------------
# Session management
# ---------------------------------------------------------------------------
def create_session(username: str, role: str) -> str:
    sid = secrets.token_hex(32)
    now = time.time()
    with sessions_lock:
        sessions[sid] = {
            "username": username,
            "role": role,
            "created_at": now,
            "expires_at": now + SESSION_TTL,
        }
    return sid


def get_session_by_id(sid: str) -> dict | None:
    if not sid:
        return None
    with sessions_lock:
        s = sessions.get(sid)
        if s and s["expires_at"] > time.time():
            return s
        if s:
            sessions.pop(sid, None)
    return None


def delete_session(sid: str):
    with sessions_lock:
        sessions.pop(sid, None)


def _cleanup_sessions_loop():
    while True:
        time.sleep(300)  # every 5 minutes
        now = time.time()
        with sessions_lock:
            expired = [k for k, v in sessions.items() if v["expires_at"] <= now]
            for k in expired:
                del sessions[k]
        # Clean up chat history for expired sessions
        for k in expired:
            clear_chat_history(k)


# Start background cleanup thread
threading.Thread(target=_cleanup_sessions_loop, daemon=True).start()


# ---------------------------------------------------------------------------
# Chat memory — per-session conversation history
# ---------------------------------------------------------------------------
chat_sessions: dict[str, list[dict]] = {}  # session_id -> messages
chat_session_meta: dict[str, dict] = {}  # session_id -> {username, created_at, last_activity, model}
chat_sessions_lock = threading.Lock()
MAX_CHAT_HISTORY = 20


def get_chat_history(session_id: str) -> list[dict]:
    with chat_sessions_lock:
        return list(chat_sessions.get(session_id, []))


def append_chat_message(session_id: str, role: str, content: str, username: str = "", model: str = ""):
    now = time.time()
    with chat_sessions_lock:
        if session_id not in chat_sessions:
            chat_sessions[session_id] = []
        chat_sessions[session_id].append({"role": role, "content": content, "timestamp": now})
        # Trim to max history (keep last MAX_CHAT_HISTORY messages)
        if len(chat_sessions[session_id]) > MAX_CHAT_HISTORY:
            chat_sessions[session_id] = chat_sessions[session_id][-MAX_CHAT_HISTORY:]
        # Update session metadata
        if session_id not in chat_session_meta:
            chat_session_meta[session_id] = {
                "username": username,
                "created_at": now,
                "last_activity": now,
                "model": model,
            }
        else:
            chat_session_meta[session_id]["last_activity"] = now
            if username:
                chat_session_meta[session_id]["username"] = username
            if model:
                chat_session_meta[session_id]["model"] = model


def clear_chat_history(session_id: str):
    with chat_sessions_lock:
        chat_sessions.pop(session_id, None)
        chat_session_meta.pop(session_id, None)


# ---------------------------------------------------------------------------
# OpenAI OAuth PKCE helpers
# ---------------------------------------------------------------------------
_openai_pkce_pending: dict[str, dict] = {}  # state -> {verifier, created_at}
_pkce_lock = threading.Lock()


def _pkce_code_verifier() -> str:
    """Generate a PKCE code verifier (64 bytes, base64url-encoded, matching Codex CLI)."""
    return secrets.token_urlsafe(64)


def _pkce_code_challenge(verifier: str) -> str:
    """SHA256 + base64url-encode the verifier."""
    digest = hashlib.sha256(verifier.encode("ascii")).digest()
    return base64.urlsafe_b64encode(digest).rstrip(b"=").decode("ascii")


_oauth_callback_server = None
_oauth_callback_lock = threading.Lock()


def _start_oauth_callback_server(state: str):
    """Spin up a temporary HTTP server on port 1455 to catch the OAuth callback.

    When OpenAI redirects to http://localhost:1455/auth/callback?code=X&state=Y,
    this server captures it and redirects the browser to our main dashboard
    with the code+state params so the JS can complete the exchange.
    """
    global _oauth_callback_server

    with _oauth_callback_lock:
        # If already running, don't start another
        if _oauth_callback_server is not None:
            try:
                _oauth_callback_server.shutdown()
            except Exception:
                pass
            _oauth_callback_server = None

    main_port = PORT

    class CallbackHandler(BaseHTTPRequestHandler):
        def log_message(self, fmt, *a):
            pass

        def do_GET(self):
            parsed = urlparse(self.path)
            qs = parse_qs(parsed.query)
            code = (qs.get("code") or [""])[0]
            cb_state = (qs.get("state") or [""])[0]
            error = (qs.get("error") or [""])[0]

            if error:
                html = f"""<!DOCTYPE html><html><body>
                <h2>OpenAI Login Failed</h2><p>Error: {error}</p>
                <p><a href="http://localhost:{main_port}/">Back to Dashboard</a></p>
                </body></html>"""
            else:
                # Redirect to main dashboard with code+state
                redirect_url = f"http://localhost:{main_port}/?code={code}&state={cb_state}"
                html = f"""<!DOCTYPE html><html><head>
                <meta http-equiv="refresh" content="0;url={redirect_url}">
                </head><body>Redirecting...</body></html>"""

            body = html.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

            # Shut down after handling the callback (in a thread to not deadlock)
            threading.Thread(target=self._shutdown, daemon=True).start()

        def _shutdown(self):
            global _oauth_callback_server
            time.sleep(1)
            with _oauth_callback_lock:
                if _oauth_callback_server is not None:
                    try:
                        _oauth_callback_server.shutdown()
                    except Exception:
                        pass
                    _oauth_callback_server = None

    def _run():
        global _oauth_callback_server
        try:
            from http.server import HTTPServer
            srv = HTTPServer(("127.0.0.1", OPENAI_CALLBACK_PORT), CallbackHandler)
            srv.timeout = 300  # 5 min max wait
            with _oauth_callback_lock:
                _oauth_callback_server = srv
            srv.serve_forever()
        except OSError:
            pass  # Port already in use

    threading.Thread(target=_run, daemon=True).start()


def refresh_openai_token() -> str | None:
    """Use refresh_token to get a new access token. Also attempts API key exchange. Returns new token or None."""
    with config_lock:
        refresh_tok = config.get("openai_refresh_token", "")
    if not refresh_tok:
        return None

    try:
        body = json.dumps({
            "grant_type": "refresh_token",
            "client_id": OPENAI_CLIENT_ID,
            "refresh_token": refresh_tok,
            "scope": "openid profile email",
        }).encode("utf-8")
        req = Request(
            OPENAI_TOKEN_URL, data=body,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        resp = urlopen(req, timeout=10)
        result = json.loads(resp.read())
        access_token = result.get("access_token", "")
        id_token = result.get("id_token", "")
        new_refresh = result.get("refresh_token", refresh_tok)
        expires_in = result.get("expires_in", 3600)

        # Try to exchange id_token for API key
        api_key = ""
        if id_token:
            try:
                exchange_body = urlencode({
                    "grant_type": "urn:ietf:params:oauth:grant-type:token-exchange",
                    "client_id": OPENAI_CLIENT_ID,
                    "requested_token": "openai-api-key",
                    "subject_token": id_token,
                    "subject_token_type": "urn:ietf:params:oauth:token-type:id_token",
                }).encode("utf-8")
                req2 = Request(
                    OPENAI_TOKEN_URL, data=exchange_body,
                    headers={"Content-Type": "application/x-www-form-urlencoded"},
                    method="POST",
                )
                resp2 = urlopen(req2, timeout=10)
                result2 = json.loads(resp2.read())
                api_key = result2.get("access_token", "")
                print(f"[oauth] Refresh token exchange succeeded, got API key: {bool(api_key)}", file=sys.stderr)
            except Exception as exc:
                print(f"[oauth] Refresh token exchange for API key failed: {exc}", file=sys.stderr)

        with config_lock:
            config["openai_oauth_token"] = access_token
            config["openai_refresh_token"] = new_refresh
            config["openai_token_expires"] = time.time() + expires_in - 60
            if api_key:
                config["openai_api_key"] = api_key
        save_config()
        return api_key if api_key else access_token
    except Exception as e:
        print(f"[oauth] Token refresh failed: {e}", file=sys.stderr)
        return None


def get_openai_bearer_token() -> str:
    """Return a valid OpenAI bearer token. Prefers API key from token exchange, then OAuth, then manual key."""
    with config_lock:
        api_key = config.get("openai_api_key", "")
        oauth_token = config.get("openai_oauth_token", "")
        expires = config.get("openai_token_expires", 0)

    # Prefer API key (obtained via OAuth token exchange) — works with all endpoints
    if api_key:
        return api_key

    # If we have a valid OAuth token, use it
    if oauth_token and time.time() < expires:
        return oauth_token

    # Try to refresh
    if config.get("openai_refresh_token"):
        new_token = refresh_openai_token()
        if new_token:
            return new_token

    return ""


# ---------------------------------------------------------------------------
# Config persistence
# ---------------------------------------------------------------------------
def load_config():
    global config
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                saved = json.load(f)
            config.update(saved)
        except Exception:
            pass


def save_config():
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[config] Error saving: {e}", file=sys.stderr)


# ---------------------------------------------------------------------------
# User management persistence
# ---------------------------------------------------------------------------
def load_users():
    global users
    if USERS_FILE.exists():
        try:
            with open(USERS_FILE, "r", encoding="utf-8") as f:
                users = json.load(f)
        except Exception:
            pass


def save_users():
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with users_lock:
            with open(USERS_FILE, "w", encoding="utf-8") as f:
                json.dump(users, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[users] Error saving: {e}", file=sys.stderr)


def touch_user(user_id: str):
    """Auto-register / update last_seen for a user on every event."""
    if not user_id or user_id == "admin":
        return
    now = time.time()
    with users_lock:
        if user_id not in users:
            users[user_id] = {
                "user_id": user_id,
                "status": "active",
                "first_seen": now,
                "last_seen": now,
                "message_count": 0,
                "error_count": 0,
                "note": "",
                "priority": "normal",
            }
        users[user_id]["last_seen"] = now


def increment_user_stat(user_id: str, field: str, amount: int = 1):
    if not user_id or user_id == "admin":
        return
    with users_lock:
        if user_id in users:
            users[user_id][field] = users[user_id].get(field, 0) + amount


def is_user_allowed(user_id: str) -> bool:
    """Check if a user is allowed to use the bot."""
    if not user_id or user_id == "admin":
        return True
    with users_lock:
        u = users.get(user_id)
        if u and u.get("status") == "blocked":
            return False
    return True


def get_user_summary(user_id: str) -> dict:
    """Get user info + recent activity from events (uses pre-indexed lookups)."""
    with users_lock:
        u = dict(users.get(user_id, {}))
    if not u:
        u = {"user_id": user_id, "status": "unknown"}

    with events_lock:
        user_events = _events_by_user.get(user_id, [])
        total = len(user_events)
        chats = sum(1 for e in user_events if e.get("type") == "chat")
        queries = sum(1 for e in user_events if e.get("type") == "query")
        errors = sum(1 for e in user_events if e.get("type") == "error")
        recent = user_events[-20:]

    u["total_events"] = total
    u["total_chats"] = chats
    u["total_queries"] = queries
    u["total_errors"] = errors
    u["recent_events"] = recent
    return u


# ---------------------------------------------------------------------------
# Event persistence
# ---------------------------------------------------------------------------
def load_events():
    global events
    if not EVENTS_FILE.exists():
        return
    loaded = []
    with open(EVENTS_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    evt = json.loads(line)
                    loaded.append(evt)
                    mod = evt.get("_moderation")
                    if mod:
                        moderation[evt["id"]] = mod
                except json.JSONDecodeError:
                    continue
    events = deque(loaded[-MAX_EVENTS:], maxlen=MAX_EVENTS)
    # Rebuild indexes
    _events_by_user.clear()
    _events_by_type.clear()
    for evt in events:
        uid = evt.get("user_id")
        if uid:
            _events_by_user[uid].append(evt)
        _events_by_type[evt.get("type", "unknown")] += 1


def rebuild_users_from_events():
    """Rebuild user stats from loaded events (on startup)."""
    for evt in events:
        uid = evt.get("user_id")
        if uid and uid != "admin":
            touch_user(uid)
            if evt.get("type") == "chat":
                increment_user_stat(uid, "message_count")
            elif evt.get("type") == "error":
                increment_user_stat(uid, "error_count")


def append_event_to_file(evt: dict):
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(EVENTS_FILE, "a", encoding="utf-8") as f:
            f.write(json.dumps(evt, ensure_ascii=False) + "\n")
    except Exception as e:
        print(f"[persist] Error writing event: {e}", file=sys.stderr)


_moderation_save_pending = False
_moderation_save_lock = threading.Lock()


def persist_moderation(event_id: str, mod: dict):
    """Update moderation on event in memory; batch-write to disk (debounced)."""
    global _moderation_save_pending
    with events_lock:
        for evt in events:
            if evt["id"] == event_id:
                evt["_moderation"] = mod
                break

    # Debounce: only rewrite file once even if many moderation updates come quickly
    with _moderation_save_lock:
        if _moderation_save_pending:
            return
        _moderation_save_pending = True

    def _flush():
        global _moderation_save_pending
        time.sleep(2)  # batch writes within 2 seconds
        try:
            DATA_DIR.mkdir(parents=True, exist_ok=True)
            with events_lock:
                snapshot = list(events)
            with open(EVENTS_FILE, "w", encoding="utf-8") as f:
                for evt in snapshot:
                    f.write(json.dumps(evt, ensure_ascii=False) + "\n")
        except Exception as e:
            print(f"[persist] Error writing moderation: {e}", file=sys.stderr)
        finally:
            with _moderation_save_lock:
                _moderation_save_pending = False

    threading.Thread(target=_flush, daemon=True).start()


def clear_all_events():
    global moderation
    with events_lock:
        events.clear()
        _events_by_user.clear()
        _events_by_type.clear()
    with moderation_lock:
        moderation = {}
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(EVENTS_FILE, "w", encoding="utf-8") as f:
            pass
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Ingest + broadcast
# ---------------------------------------------------------------------------
def ingest_event(evt: dict) -> dict:
    if "id" not in evt:
        evt["id"] = str(uuid.uuid4())
    if "timestamp" not in evt:
        evt["timestamp"] = time.time()
    if "type" not in evt:
        evt["type"] = "system"

    # Track user
    uid = evt.get("user_id")
    if uid:
        touch_user(uid)
        if evt["type"] == "chat":
            increment_user_stat(uid, "message_count")
        elif evt["type"] == "error":
            increment_user_stat(uid, "error_count")

    with events_lock:
        # If deque is full, the oldest event is auto-evicted — update index
        if len(events) == events.maxlen:
            old = events[0]
            old_uid = old.get("user_id")
            if old_uid and old_uid in _events_by_user:
                user_list = _events_by_user[old_uid]
                if user_list and user_list[0] is old:
                    user_list.pop(0)
            old_type = old.get("type", "unknown")
            if _events_by_type[old_type] > 0:
                _events_by_type[old_type] -= 1
        events.append(evt)
        # Update indexes
        if uid:
            _events_by_user[uid].append(evt)
        _events_by_type[evt.get("type", "system")] += 1

    append_event_to_file(evt)
    broadcast_event(evt)

    # Save users periodically (every 10th event to avoid thrashing)
    if len(events) % 10 == 0:
        save_users()

    return evt


# ---------------------------------------------------------------------------
# LLM providers
# ---------------------------------------------------------------------------
def call_ollama(messages: list[dict], model: str, temperature: float) -> tuple[str, float, dict]:
    with config_lock:
        ollama_url = config["ollama_url"]

    body: dict = {"model": model, "messages": messages, "stream": False}
    if temperature is not None:
        body["options"] = {"temperature": temperature}

    start = time.time()
    data = json.dumps(body).encode("utf-8")
    req = Request(
        f"{ollama_url}/api/chat", data=data,
        headers={"Content-Type": "application/json"}, method="POST",
    )
    resp = urlopen(req, timeout=600)
    result = json.loads(resp.read())
    answer = result.get("message", {}).get("content", "")
    # Ollama returns eval_count (output tokens) and prompt_eval_count (input tokens)
    token_usage = {
        "prompt_tokens": result.get("prompt_eval_count", 0),
        "completion_tokens": result.get("eval_count", 0),
        "total_tokens": result.get("prompt_eval_count", 0) + result.get("eval_count", 0),
    }
    duration_ms = (time.time() - start) * 1000
    return answer, round(duration_ms, 1), token_usage


CHATGPT_BACKEND_URL = "https://chatgpt.com/backend-api/codex"

# Codex models that work with ChatGPT backend (NOT standard API models)
CHATGPT_CODEX_MODELS = [
    {"name": "gpt-5.3-codex", "size": 0},
    {"name": "gpt-5.2-codex", "size": 0},
    {"name": "gpt-5-codex", "size": 0},
    {"name": "gpt-5-codex-mini", "size": 0},
    {"name": "gpt-5.2", "size": 0},
    {"name": "gpt-5", "size": 0},
    {"name": "gpt-4o", "size": 0},
]

# Default model for ChatGPT backend
CHATGPT_DEFAULT_MODEL = "gpt-4o"


def _decode_jwt_payload(token: str) -> dict:
    """Decode JWT payload without verification (just to extract claims)."""
    try:
        parts = token.split(".")
        if len(parts) < 2:
            return {}
        payload = parts[1]
        # Add padding
        payload += "=" * (4 - len(payload) % 4)
        decoded = base64.urlsafe_b64decode(payload)
        return json.loads(decoded)
    except Exception:
        return {}


def _get_chatgpt_account_id() -> str:
    """Extract ChatGPT account ID from stored OAuth token's JWT claims."""
    with config_lock:
        token = config.get("openai_oauth_token", "")
        # Check if explicitly stored
        stored_id = config.get("openai_chatgpt_account_id", "")
    if stored_id:
        return stored_id
    if not token:
        return ""
    claims = _decode_jwt_payload(token)
    account_id = ""
    # Check organizations list
    orgs = claims.get("organizations", [])
    if orgs and isinstance(orgs, list):
        if isinstance(orgs[0], dict):
            account_id = orgs[0].get("id", "")
        elif isinstance(orgs[0], str):
            account_id = orgs[0]
    # Check nested auth claim (ChatGPT JWT stores it as chatgpt_account_id)
    if not account_id:
        auth_claim = claims.get("https://api.openai.com/auth", {})
        if isinstance(auth_claim, dict):
            account_id = (auth_claim.get("chatgpt_account_id", "") or
                          auth_claim.get("account_id", "") or
                          auth_claim.get("organization_id", ""))
    # Check top-level claims
    if not account_id:
        account_id = claims.get("account_id", "") or claims.get("org_id", "") or claims.get("sub", "")
    # Cache it
    if account_id:
        with config_lock:
            config["openai_chatgpt_account_id"] = account_id
    return account_id


def _using_oauth_token() -> bool:
    """Check if we're using a raw OAuth token (no API key from token exchange)."""
    with config_lock:
        has_api_key = bool(config.get("openai_api_key"))
        has_oauth = bool(config.get("openai_oauth_token"))
    # Use ChatGPT backend if we have OAuth but no proper API key
    return has_oauth and not has_api_key


def _call_openai_chat_completions(base_url: str, bearer: str, messages: list[dict], model: str, temperature: float) -> tuple[str, dict]:
    """Call OpenAI Chat Completions API. Returns (answer, usage_dict)."""
    if not model:
        model = "gpt-4o-mini"
    body = {"model": model, "messages": messages}
    if temperature is not None:
        body["temperature"] = temperature
    data = json.dumps(body).encode("utf-8")
    print(f"[openai] Chat Completions request: model={model}, messages={len(messages)}", file=sys.stderr)
    req = Request(
        f"{base_url}/chat/completions", data=data,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {bearer}",
        },
        method="POST",
    )
    try:
        resp = urlopen(req, timeout=120)
    except Exception as e:
        if hasattr(e, 'read'):
            try:
                err_body = e.read().decode()
                print(f"[openai] Chat Completions error body: {err_body}", file=sys.stderr)
            except Exception:
                pass
        raise
    result = json.loads(resp.read())
    usage = result.get("usage", {})
    token_usage = {
        "prompt_tokens": usage.get("prompt_tokens", 0),
        "completion_tokens": usage.get("completion_tokens", 0),
        "total_tokens": usage.get("total_tokens", 0),
    }
    return result["choices"][0]["message"]["content"], token_usage


def _call_openai_responses(base_url: str, bearer: str, messages: list[dict], model: str, temperature: float,
                           is_chatgpt_backend: bool = False) -> tuple[str, dict]:
    """Call OpenAI Responses API. Returns (answer, usage_dict)."""
    if not model:
        model = CHATGPT_DEFAULT_MODEL if is_chatgpt_backend else "gpt-4o-mini"
    system_text = ""
    input_parts = []
    for m in messages:
        if m["role"] == "system":
            system_text = m["content"]
        else:
            input_parts.append(m)

    body = {"model": model}
    # ChatGPT backend REQUIRES instructions field (even if empty)
    body["instructions"] = system_text or "You are a helpful assistant. Antworte auf Deutsch."

    if is_chatgpt_backend:
        # ChatGPT backend needs explicit content type format
        # User messages use "input_text", assistant messages use "output_text"
        body["input"] = [
            {
                "role": m["role"],
                "type": "message",
                "content": [{"type": "output_text" if m["role"] == "assistant" else "input_text", "text": m["content"]}],
            }
            for m in input_parts
        ]
    else:
        body["input"] = [{"role": m["role"], "content": m["content"]} for m in input_parts]

    if temperature is not None and not is_chatgpt_backend:
        body["temperature"] = temperature

    # ChatGPT backend: requires stream=true, don't store
    if is_chatgpt_backend:
        body["store"] = False
        body["stream"] = True

    data = json.dumps(body).encode("utf-8")
    url = f"{base_url}/responses"
    print(f"[openai] Responses API request: url={url}, model={model}, input_parts={len(input_parts)}, chatgpt_backend={is_chatgpt_backend}", file=sys.stderr)
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {bearer}",
    }

    # ChatGPT backend requires account ID header
    if is_chatgpt_backend:
        account_id = _get_chatgpt_account_id()
        if account_id:
            headers["ChatGPT-Account-ID"] = account_id
            print(f"[openai] Using ChatGPT-Account-ID: {account_id[:8]}...", file=sys.stderr)

    req = Request(url, data=data, headers=headers, method="POST")
    try:
        resp = urlopen(req, timeout=120)
    except Exception as e:
        if hasattr(e, 'read'):
            try:
                err_body = e.read().decode()
                print(f"[openai] Responses API error body: {err_body}", file=sys.stderr)
            except Exception:
                pass
        raise

    token_usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}

    if is_chatgpt_backend:
        # Parse SSE stream: collect text deltas from response.output_text.delta events
        answer = ""
        for raw_line in resp:
            line = raw_line.decode("utf-8", errors="replace").strip()
            if not line or not line.startswith("data: "):
                continue
            payload = line[6:]  # strip "data: "
            if payload == "[DONE]":
                break
            try:
                evt = json.loads(payload)
                evt_type = evt.get("type", "")
                if evt_type == "response.output_text.delta":
                    answer += evt.get("delta", "")
                elif evt_type == "response.completed":
                    # Final event — extract usage + full text if we missed deltas
                    resp_obj = evt.get("response", {})
                    usage = resp_obj.get("usage", {})
                    token_usage["prompt_tokens"] = usage.get("input_tokens", 0)
                    token_usage["completion_tokens"] = usage.get("output_tokens", 0)
                    token_usage["total_tokens"] = usage.get("total_tokens",
                                                            token_usage["prompt_tokens"] + token_usage["completion_tokens"])
                    if not answer:
                        for item in resp_obj.get("output", []):
                            if item.get("type") == "message":
                                for c in item.get("content", []):
                                    if c.get("type") == "output_text":
                                        answer += c.get("text", "")
                elif evt_type == "error":
                    err_msg = evt.get("message", "") or evt.get("error", {}).get("message", "")
                    raise ValueError(f"Stream error: {err_msg}")
            except json.JSONDecodeError:
                continue
        if not answer:
            raise ValueError("No text received from ChatGPT backend stream")
        return answer, token_usage
    else:
        result = json.loads(resp.read())
        answer = ""
        for item in result.get("output", []):
            if item.get("type") == "message":
                for c in item.get("content", []):
                    if c.get("type") == "output_text":
                        answer += c.get("text", "")
        if not answer:
            answer = result.get("output_text", "") or str(result)
        usage = result.get("usage", {})
        token_usage["prompt_tokens"] = usage.get("input_tokens", usage.get("prompt_tokens", 0))
        token_usage["completion_tokens"] = usage.get("output_tokens", usage.get("completion_tokens", 0))
        token_usage["total_tokens"] = usage.get("total_tokens",
                                                 token_usage["prompt_tokens"] + token_usage["completion_tokens"])
        return answer, token_usage


def call_openai(messages: list[dict], model: str, temperature: float) -> tuple[str, float, dict]:
    with config_lock:
        base_url = config.get("openai_base_url", "https://api.openai.com/v1").rstrip("/")

    bearer = get_openai_bearer_token()
    if not bearer:
        raise ValueError("OpenAI API key not configured — use Login with OpenAI or set an API key in Settings")

    use_oauth = _using_oauth_token()

    start = time.time()
    token_usage = {}

    if use_oauth:
        # OAuth token (ChatGPT login): must use chatgpt.com backend, NOT api.openai.com
        print(f"[openai] Using ChatGPT backend (OAuth token), model={model}", file=sys.stderr)
        try:
            answer, token_usage = _call_openai_responses(CHATGPT_BACKEND_URL, bearer, messages, model, temperature,
                                                         is_chatgpt_backend=True)
        except Exception as e1:
            print(f"[openai] ChatGPT backend failed: {e1}", file=sys.stderr)
            if "429" in str(e1):
                raise ValueError("OpenAI rate limit reached — your account usage limit may be exceeded. Please wait or check your OpenAI plan.")
            raise ValueError(f"OpenAI API call failed. Error: {e1}")
    else:
        # API key: use standard api.openai.com with Chat Completions
        answer, token_usage = _call_openai_chat_completions(base_url, bearer, messages, model, temperature)

    duration_ms = (time.time() - start) * 1000
    return answer, round(duration_ms, 1), token_usage


def call_anthropic(messages: list[dict], model: str, temperature: float) -> tuple[str, float, dict]:
    with config_lock:
        api_key = config["anthropic_api_key"]

    if not api_key:
        raise ValueError("Anthropic API key not configured")

    # Anthropic API requires system prompt in a separate field
    system_prompt = ""
    chat_messages = []
    for m in messages:
        if m["role"] == "system":
            system_prompt = m["content"]
        else:
            chat_messages.append(m)

    body: dict = {
        "model": model,
        "max_tokens": 4096,
        "messages": chat_messages,
    }
    if system_prompt:
        body["system"] = system_prompt
    if temperature is not None:
        body["temperature"] = temperature

    start = time.time()
    data = json.dumps(body).encode("utf-8")
    req = Request(
        "https://api.anthropic.com/v1/messages", data=data,
        headers={
            "Content-Type": "application/json",
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
        },
        method="POST",
    )
    resp = urlopen(req, timeout=120)
    result = json.loads(resp.read())
    answer = result["content"][0]["text"]
    usage = result.get("usage", {})
    token_usage = {
        "prompt_tokens": usage.get("input_tokens", 0),
        "completion_tokens": usage.get("output_tokens", 0),
        "total_tokens": usage.get("input_tokens", 0) + usage.get("output_tokens", 0),
    }
    duration_ms = (time.time() - start) * 1000
    return answer, round(duration_ms, 1), token_usage


def call_openrouter(messages: list[dict], model: str, temperature: float) -> tuple[str, float, dict]:
    with config_lock:
        api_key = config["openrouter_api_key"]

    if not api_key:
        raise ValueError("OpenRouter API key not configured — use Login with OpenRouter in Settings")

    body = {"model": model, "messages": messages}
    if temperature is not None:
        body["temperature"] = temperature

    start = time.time()
    data = json.dumps(body).encode("utf-8")
    req = Request(
        "https://openrouter.ai/api/v1/chat/completions", data=data,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
            "HTTP-Referer": "http://localhost:7779",
            "X-Title": "WaWi Chatbot Monitor",
        },
        method="POST",
    )
    resp = urlopen(req, timeout=120)
    result = json.loads(resp.read())
    answer = result["choices"][0]["message"]["content"]
    usage = result.get("usage", {})
    token_usage = {
        "prompt_tokens": usage.get("prompt_tokens", 0),
        "completion_tokens": usage.get("completion_tokens", 0),
        "total_tokens": usage.get("total_tokens", 0),
    }
    duration_ms = (time.time() - start) * 1000
    return answer, round(duration_ms, 1), token_usage


def call_llm(messages: list[dict], model: str, temperature: float) -> tuple[str, float, dict]:
    """Dispatch to the configured LLM provider. Returns (answer, duration_ms, token_usage)."""
    with config_lock:
        provider = config.get("provider", "ollama")
    if provider == "openai":
        return call_openai(messages, model, temperature)
    elif provider == "anthropic":
        return call_anthropic(messages, model, temperature)
    elif provider == "openrouter":
        return call_openrouter(messages, model, temperature)
    else:
        return call_ollama(messages, model, temperature)


# ---------------------------------------------------------------------------
# Token usage tracking
# ---------------------------------------------------------------------------
_token_usage_lock = threading.Lock()
_token_usage: dict = {
    "session_total": {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0, "requests": 0},
    "by_user": {},  # username -> {prompt_tokens, completion_tokens, total_tokens, requests}
}


def track_token_usage(username: str, usage: dict):
    """Accumulate token usage for a user and the session total."""
    if not usage or not usage.get("total_tokens"):
        return
    with _token_usage_lock:
        for key in ("prompt_tokens", "completion_tokens", "total_tokens"):
            _token_usage["session_total"][key] += usage.get(key, 0)
        _token_usage["session_total"]["requests"] += 1
        if username not in _token_usage["by_user"]:
            _token_usage["by_user"][username] = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0, "requests": 0}
        for key in ("prompt_tokens", "completion_tokens", "total_tokens"):
            _token_usage["by_user"][username][key] += usage.get(key, 0)
        _token_usage["by_user"][username]["requests"] += 1


def get_token_usage(username: str = None) -> dict:
    """Get token usage stats. If username given, include user-specific stats."""
    with _token_usage_lock:
        result = {"session_total": dict(_token_usage["session_total"])}
        if username and username in _token_usage["by_user"]:
            result["user"] = dict(_token_usage["by_user"][username])
        else:
            result["user"] = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0, "requests": 0}
        result["all_users"] = {k: dict(v) for k, v in _token_usage["by_user"].items()}
    return result


# ---------------------------------------------------------------------------
# Agent Mode — tool registry, handlers, and loop
# ---------------------------------------------------------------------------
AGENT_TOOLS = [
    {
        "name": "query_database",
        "description": "Execute a read-only SQL query against the configured database. Use this when you need to look up data, records, or statistics.",
        "parameters": {
            "type": "object",
            "properties": {
                "sql": {"type": "string", "description": "The SQL SELECT query to execute"}
            },
            "required": ["sql"]
        },
        "privileged": True,
        "requires": "db_enabled"
    },
    {
        "name": "search_knowledge",
        "description": "Search the knowledge base (RAG) for relevant information. Use this for company-specific or uploaded document questions.",
        "parameters": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "The search query"}
            },
            "required": ["query"]
        },
        "privileged": True,
        "requires": "rag_enabled"
    },
    {
        "name": "web_search",
        "description": "Search the web using DuckDuckGo. Use for current events, general knowledge, or anything not in the database/knowledge base.",
        "parameters": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "The search query"}
            },
            "required": ["query"]
        },
        "privileged": False,
        "requires": "agent_web_search"
    },
    {
        "name": "calculate",
        "description": "Evaluate a mathematical expression safely. Supports +, -, *, /, **, %, sqrt, abs, round, min, max, sum.",
        "parameters": {
            "type": "object",
            "properties": {
                "expression": {"type": "string", "description": "Math expression, e.g. '(100 * 1.19) + 50'"}
            },
            "required": ["expression"]
        },
        "privileged": False,
        "requires": None
    },
    {
        "name": "get_datetime",
        "description": "Get the current date and time.",
        "parameters": {
            "type": "object",
            "properties": {}
        },
        "privileged": False,
        "requires": None
    },
    {
        "name": "list_tables",
        "description": "List all tables and their columns in the connected database.",
        "parameters": {
            "type": "object",
            "properties": {}
        },
        "privileged": True,
        "requires": "db_enabled"
    }
]


def _get_available_tools(role: str) -> list:
    """Filter AGENT_TOOLS based on user role and enabled features."""
    with config_lock:
        cfg_snapshot = dict(config)
    tools = []
    for t in AGENT_TOOLS:
        # Check if required feature is enabled
        req = t.get("requires")
        if req and not cfg_snapshot.get(req, False):
            continue
        # Check privilege: only admin/mitarbeiter get privileged tools
        if t.get("privileged") and role not in ("admin", "mitarbeiter"):
            continue
        tools.append(t)
    return tools


def _execute_tool(name: str, arguments: dict) -> str:
    """Execute a tool by name and return the result as a string."""
    try:
        if name == "query_database":
            sql = arguments.get("sql", "")
            result = execute_db_query(sql)
            if result.get("error"):
                return f"Error: {result['error']}"
            cols = result.get("columns", [])
            rows = result.get("rows", [])
            if not rows:
                return "Query returned 0 rows."
            # Format as markdown table
            header = " | ".join(cols)
            sep = " | ".join("---" for _ in cols)
            data = "\n".join(" | ".join(str(r.get(c, "")) for c in cols) for r in rows[:50])
            return f"{result['row_count']} rows returned:\n\n{header}\n{sep}\n{data}"

        elif name == "search_knowledge":
            query = arguments.get("query", "")
            chunks = search_embeddings(query)
            if not chunks:
                return "No relevant results found in the knowledge base."
            parts = []
            for i, c in enumerate(chunks, 1):
                parts.append(f"{i}. [Score: {c['similarity']}] {c['content'][:500]}")
            return "\n\n".join(parts)

        elif name == "web_search":
            query = arguments.get("query", "")
            return _web_search_ddg(query)

        elif name == "calculate":
            expr = arguments.get("expression", "")
            return _safe_math_eval(expr)

        elif name == "get_datetime":
            now = datetime.now()
            return now.strftime("%A, %d. %B %Y, %H:%M:%S Uhr")

        elif name == "list_tables":
            schema = get_db_schema()
            return schema if schema else "No database schema available."

        else:
            return f"Unknown tool: {name}"
    except Exception as e:
        return f"Tool error ({name}): {e}"


def _web_search_ddg(query: str, max_results: int = 5) -> str:
    """Search DuckDuckGo HTML and extract results."""
    try:
        from urllib.parse import quote_plus
        url = f"https://html.duckduckgo.com/html/?q={quote_plus(query)}"
        req = Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        })
        resp = urlopen(req, timeout=10)
        html = resp.read().decode("utf-8", errors="replace")

        # Parse results
        results = []
        # Find result links
        link_pattern = re.compile(
            r'<a\s+rel="nofollow"\s+class="result__a"\s+href="([^"]+)"[^>]*>(.*?)</a>',
            re.DOTALL
        )
        snippet_pattern = re.compile(
            r'<a\s+class="result__snippet"[^>]*>(.*?)</a>',
            re.DOTALL
        )

        links = link_pattern.findall(html)
        snippets = snippet_pattern.findall(html)

        for i in range(min(len(links), max_results)):
            href, title = links[i]
            title = re.sub(r'<[^>]+>', '', title).strip()
            snippet = re.sub(r'<[^>]+>', '', snippets[i]).strip() if i < len(snippets) else ""
            if title:
                results.append(f"{i+1}. [{title}]({href})\n   {snippet}")

        if not results:
            return f"No web results found for: {query}"
        return "\n\n".join(results)
    except Exception as e:
        return f"Web search error: {e}"


def _safe_math_eval(expr: str) -> str:
    """Safely evaluate a math expression using AST parsing."""
    # Replace common German/locale decimal comma: "2.499,99" -> "2499.99"
    # But only do simple comma->dot replacement for single numbers like "2499,99"
    expr = re.sub(r'(\d),(\d)', r'\1.\2', expr)

    allowed_names = {
        "sqrt": math.sqrt, "abs": abs, "round": round,
        "min": min, "max": max, "sum": sum, "pow": pow,
        "int": int, "float": float, "pi": math.pi, "e": math.e,
    }

    try:
        tree = ast.parse(expr, mode='eval')
    except SyntaxError as e:
        return f"Syntax error: {e}"

    # Validate AST nodes
    for node in ast.walk(tree):
        if isinstance(node, (ast.Expression, ast.BinOp, ast.UnaryOp, ast.Constant,
                             ast.Call, ast.Name, ast.Load, ast.Add, ast.Sub,
                             ast.Mult, ast.Div, ast.Pow, ast.Mod, ast.FloorDiv,
                             ast.USub, ast.UAdd, ast.List, ast.Tuple)):
            continue
        # Allow Num for older Python versions
        if hasattr(ast, 'Num') and isinstance(node, ast.Num):
            continue
        if isinstance(node, ast.Name):
            if node.id not in allowed_names:
                return f"Not allowed: {node.id}"
        elif isinstance(node, ast.Call):
            if isinstance(node.func, ast.Name) and node.func.id in allowed_names:
                continue
            return f"Function not allowed"
        elif isinstance(node, ast.Attribute):
            return "Attribute access not allowed"

    # Check all Name references are allowed
    for node in ast.walk(tree):
        if isinstance(node, ast.Name) and node.id not in allowed_names:
            return f"Unknown variable: {node.id}"

    try:
        result = eval(compile(tree, "<math>", "eval"), {"__builtins__": {}}, allowed_names)
        # Format nicely
        if isinstance(result, float):
            if result == int(result) and abs(result) < 1e15:
                return str(int(result))
            return f"{result:.6g}"
        return str(result)
    except Exception as e:
        return f"Calculation error: {e}"


# --- Provider-specific tool format converters ---

def _tools_to_openai_format(tools: list) -> list:
    """Convert AGENT_TOOLS to OpenAI function calling format."""
    return [{"type": "function", "function": {"name": t["name"], "description": t["description"], "parameters": t["parameters"]}} for t in tools]


def _tools_to_anthropic_format(tools: list) -> list:
    """Convert AGENT_TOOLS to Anthropic tool use format."""
    return [{"name": t["name"], "description": t["description"], "input_schema": t["parameters"]} for t in tools]


def _supports_native_tools() -> bool:
    """Check if current provider+config supports native function calling."""
    with config_lock:
        provider = config.get("provider", "ollama")
    if provider == "anthropic":
        return True
    if provider == "openai":
        with config_lock:
            return bool(config.get("openai_api_key"))
    if provider == "openrouter":
        return True
    if provider == "ollama":
        return True
    return False


# --- Provider-specific LLM+tools callers ---

def _call_ollama_with_tools(messages: list, model: str, temperature: float, tools: list) -> tuple:
    """Call Ollama with tool support. Returns (text_or_none, tool_calls_or_none, usage)."""
    with config_lock:
        ollama_url = config["ollama_url"]

    body = {"model": model, "messages": messages, "stream": False,
            "tools": _tools_to_openai_format(tools)}
    if temperature is not None:
        body["options"] = {"temperature": temperature}

    data = json.dumps(body).encode("utf-8")
    req = Request(
        f"{ollama_url}/api/chat", data=data,
        headers={"Content-Type": "application/json"}, method="POST",
    )
    resp = urlopen(req, timeout=600)
    result = json.loads(resp.read())
    usage = {
        "prompt_tokens": result.get("prompt_eval_count", 0),
        "completion_tokens": result.get("eval_count", 0),
        "total_tokens": result.get("prompt_eval_count", 0) + result.get("eval_count", 0),
    }

    msg = result.get("message", {})
    tool_calls_raw = msg.get("tool_calls")
    if tool_calls_raw:
        parsed = []
        for tc in tool_calls_raw:
            fn = tc.get("function", {})
            parsed.append({"name": fn.get("name", ""), "arguments": fn.get("arguments", {})})
        return None, parsed, usage
    return msg.get("content", ""), None, usage


def _call_openai_with_tools(messages: list, model: str, temperature: float, tools: list) -> tuple:
    """Call OpenAI Chat Completions with tool support. Returns (text_or_none, tool_calls_or_none, usage)."""
    with config_lock:
        base_url = config.get("openai_base_url", "https://api.openai.com/v1").rstrip("/")
    bearer = get_openai_bearer_token()
    if not bearer:
        raise ValueError("OpenAI API key not configured")

    body = {"model": model or "gpt-4o-mini", "messages": messages,
            "tools": _tools_to_openai_format(tools)}
    if temperature is not None:
        body["temperature"] = temperature

    data = json.dumps(body).encode("utf-8")
    req = Request(
        f"{base_url}/chat/completions", data=data,
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {bearer}"},
        method="POST",
    )
    resp = urlopen(req, timeout=120)
    result = json.loads(resp.read())
    usage_raw = result.get("usage", {})
    usage = {
        "prompt_tokens": usage_raw.get("prompt_tokens", 0),
        "completion_tokens": usage_raw.get("completion_tokens", 0),
        "total_tokens": usage_raw.get("total_tokens", 0),
    }

    choice = result["choices"][0]["message"]
    tool_calls_raw = choice.get("tool_calls")
    if tool_calls_raw:
        parsed = []
        for tc in tool_calls_raw:
            fn = tc.get("function", {})
            args = fn.get("arguments", "{}")
            if isinstance(args, str):
                try:
                    args = json.loads(args)
                except json.JSONDecodeError:
                    args = {}
            parsed.append({"name": fn.get("name", ""), "arguments": args, "id": tc.get("id", "")})
        return None, parsed, usage
    return choice.get("content", ""), None, usage


def _call_anthropic_with_tools(messages: list, model: str, temperature: float, tools: list) -> tuple:
    """Call Anthropic with tool support. Returns (text_or_none, tool_calls_or_none, usage)."""
    with config_lock:
        api_key = config["anthropic_api_key"]
    if not api_key:
        raise ValueError("Anthropic API key not configured")

    system_prompt = ""
    chat_messages = []
    for m in messages:
        if m["role"] == "system":
            system_prompt += m["content"] + "\n"
        else:
            chat_messages.append(m)

    body = {
        "model": model, "max_tokens": 4096,
        "messages": chat_messages,
        "tools": _tools_to_anthropic_format(tools),
    }
    if system_prompt:
        body["system"] = system_prompt.strip()
    if temperature is not None:
        body["temperature"] = temperature

    data = json.dumps(body).encode("utf-8")
    req = Request(
        "https://api.anthropic.com/v1/messages", data=data,
        headers={
            "Content-Type": "application/json",
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
        },
        method="POST",
    )
    resp = urlopen(req, timeout=120)
    result = json.loads(resp.read())
    usage_raw = result.get("usage", {})
    usage = {
        "prompt_tokens": usage_raw.get("input_tokens", 0),
        "completion_tokens": usage_raw.get("output_tokens", 0),
        "total_tokens": usage_raw.get("input_tokens", 0) + usage_raw.get("output_tokens", 0),
    }

    # Anthropic returns content blocks
    text_parts = []
    tool_calls = []
    for block in result.get("content", []):
        if block["type"] == "text":
            text_parts.append(block["text"])
        elif block["type"] == "tool_use":
            tool_calls.append({
                "name": block["name"],
                "arguments": block.get("input", {}),
                "id": block.get("id", ""),
            })

    if tool_calls:
        return " ".join(text_parts) if text_parts else None, tool_calls, usage
    return " ".join(text_parts) if text_parts else "", None, usage


def _call_openrouter_with_tools(messages: list, model: str, temperature: float, tools: list) -> tuple:
    """Call OpenRouter with tool support. Returns (text_or_none, tool_calls_or_none, usage)."""
    with config_lock:
        api_key = config["openrouter_api_key"]
    if not api_key:
        raise ValueError("OpenRouter API key not configured")

    body = {"model": model, "messages": messages,
            "tools": _tools_to_openai_format(tools)}
    if temperature is not None:
        body["temperature"] = temperature

    data = json.dumps(body).encode("utf-8")
    req = Request(
        "https://openrouter.ai/api/v1/chat/completions", data=data,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
            "HTTP-Referer": "http://localhost:7779",
            "X-Title": "WaWi Chatbot Monitor",
        },
        method="POST",
    )
    resp = urlopen(req, timeout=120)
    result = json.loads(resp.read())
    usage_raw = result.get("usage", {})
    usage = {
        "prompt_tokens": usage_raw.get("prompt_tokens", 0),
        "completion_tokens": usage_raw.get("completion_tokens", 0),
        "total_tokens": usage_raw.get("total_tokens", 0),
    }

    choice = result["choices"][0]["message"]
    tool_calls_raw = choice.get("tool_calls")
    if tool_calls_raw:
        parsed = []
        for tc in tool_calls_raw:
            fn = tc.get("function", {})
            args = fn.get("arguments", "{}")
            if isinstance(args, str):
                try:
                    args = json.loads(args)
                except json.JSONDecodeError:
                    args = {}
            parsed.append({"name": fn.get("name", ""), "arguments": args, "id": tc.get("id", "")})
        return None, parsed, usage
    return choice.get("content", ""), None, usage


def _call_with_text_fallback(messages: list, model: str, temperature: float, tools: list) -> tuple:
    """Text-based tool calling fallback for providers without native tool support."""
    # Build tool description text
    tool_desc_parts = ["You have access to the following tools:\n"]
    for t in tools:
        params_desc = ""
        props = t["parameters"].get("properties", {})
        if props:
            param_strs = []
            for pname, pinfo in props.items():
                param_strs.append(f'  - {pname}: {pinfo.get("description", "")}')
            params_desc = "\n".join(param_strs)
        tool_desc_parts.append(f"**{t['name']}**: {t['description']}")
        if params_desc:
            tool_desc_parts.append(f"  Parameters:\n{params_desc}")

    tool_desc_parts.append(
        '\nTo use a tool, respond with:\n<tool_call>{"name": "tool_name", "arguments": {"param": "value"}}</tool_call>\n'
        "You may use multiple tool calls in one response. After tool results are provided, continue reasoning."
    )
    tool_instruction = "\n".join(tool_desc_parts)

    # Inject into system message
    augmented = []
    system_injected = False
    for m in messages:
        if m["role"] == "system" and not system_injected:
            augmented.append({"role": "system", "content": m["content"] + "\n\n" + tool_instruction})
            system_injected = True
        else:
            augmented.append(m)
    if not system_injected:
        augmented.insert(0, {"role": "system", "content": tool_instruction})

    answer, duration_ms, usage = call_llm(augmented, model, temperature)

    # Parse <tool_call> blocks
    tc_pattern = re.compile(r'<tool_call>(.*?)</tool_call>', re.DOTALL)
    matches = tc_pattern.findall(answer)
    if matches:
        parsed = []
        for m in matches:
            try:
                obj = json.loads(m.strip())
                parsed.append({"name": obj.get("name", ""), "arguments": obj.get("arguments", {})})
            except json.JSONDecodeError:
                continue
        # Strip tool calls from text
        clean_text = tc_pattern.sub("", answer).strip()
        if parsed:
            return clean_text if clean_text else None, parsed, usage
    return answer, None, usage


# --- Agent tool execution helper ---

def _run_tool_and_log(tc: dict, step: int, tool_log: list, run_id: str, user_id: str) -> str:
    """Execute a single tool call, log it, broadcast SSE event. Returns result string."""
    t0 = time.time()
    result_str = _execute_tool(tc["name"], tc["arguments"])
    tool_ms = (time.time() - t0) * 1000
    tool_log.append({"tool": tc["name"], "input": tc["arguments"], "output": result_str, "step": step})
    print(f"[agent]   Tool: {tc['name']}({tc['arguments']}) -> {result_str[:200]}", file=sys.stderr)
    broadcast_event({"type": "_agent_step", "run_id": run_id, "user_id": user_id,
        "step": step, "tool": tc["name"], "input": tc["arguments"],
        "output": result_str[:2000], "duration_ms": round(tool_ms, 1), "timestamp": time.time()})
    return result_str


# --- Core agent loop ---

def run_agent_loop(messages: list, model: str, temperature: float,
                   available_tools: list, max_steps: int = 8,
                   user_id: str = "", run_id: str = "") -> tuple:
    """Run the agent loop. Returns (final_answer, tool_log, total_usage)."""
    tool_log = []
    total_usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
    use_native = _supports_native_tools()

    with config_lock:
        provider = config.get("provider", "ollama")

    # Working copy of messages
    working_messages = list(messages)

    loop_start = time.time()
    broadcast_event({"type": "_agent_start", "run_id": run_id, "user_id": user_id,
        "question": messages[-1].get("content", "") if messages else "",
        "available_tools": [t["name"] for t in available_tools],
        "max_steps": max_steps, "model": model, "timestamp": time.time()})

    for step in range(max_steps):
        print(f"[agent] Step {step+1}/{max_steps}, provider={provider}, native={use_native}", file=sys.stderr)
        broadcast_event({"type": "_agent_thinking", "run_id": run_id, "user_id": user_id,
            "step": step + 1, "timestamp": time.time()})

        try:
            if use_native:
                if provider == "ollama":
                    text, tool_calls, usage = _call_ollama_with_tools(working_messages, model, temperature, available_tools)
                elif provider == "openai":
                    text, tool_calls, usage = _call_openai_with_tools(working_messages, model, temperature, available_tools)
                elif provider == "anthropic":
                    text, tool_calls, usage = _call_anthropic_with_tools(working_messages, model, temperature, available_tools)
                elif provider == "openrouter":
                    text, tool_calls, usage = _call_openrouter_with_tools(working_messages, model, temperature, available_tools)
                else:
                    text, tool_calls, usage = _call_with_text_fallback(working_messages, model, temperature, available_tools)
            else:
                text, tool_calls, usage = _call_with_text_fallback(working_messages, model, temperature, available_tools)
        except Exception as e:
            print(f"[agent] LLM call error at step {step+1}: {e}", file=sys.stderr)
            return f"Agent error: {e}", tool_log, total_usage

        # Accumulate usage
        for k in total_usage:
            total_usage[k] += usage.get(k, 0)

        # No tool calls — LLM is done
        if not tool_calls:
            final = text or ""
            if step == 0 and not final:
                final = "(No response from model)"
            broadcast_event({"type": "_agent_done", "run_id": run_id, "user_id": user_id,
                "answer": final[:500], "total_steps": len(tool_log),
                "duration_ms": round((time.time() - loop_start) * 1000, 1), "timestamp": time.time()})
            return final, tool_log, total_usage

        # Execute tool calls and build follow-up messages
        if use_native and provider == "anthropic":
            # Anthropic: assistant message has content blocks, tool results go as user message
            assistant_content = []
            if text:
                assistant_content.append({"type": "text", "text": text})
            for tc in tool_calls:
                assistant_content.append({
                    "type": "tool_use",
                    "id": tc.get("id", f"tool_{step}_{tc['name']}"),
                    "name": tc["name"],
                    "input": tc["arguments"],
                })
            working_messages.append({"role": "assistant", "content": assistant_content})

            tool_results_content = []
            for tc in tool_calls:
                result_str = _run_tool_and_log(tc, step + 1, tool_log, run_id, user_id)
                tool_results_content.append({
                    "type": "tool_result",
                    "tool_use_id": tc.get("id", f"tool_{step}_{tc['name']}"),
                    "content": result_str,
                })
            working_messages.append({"role": "user", "content": tool_results_content})

        elif use_native and provider in ("openai", "openrouter"):
            # OpenAI/OpenRouter: assistant message with tool_calls, then tool role messages
            assistant_msg = {"role": "assistant", "content": text or ""}
            assistant_msg["tool_calls"] = []
            for i, tc in enumerate(tool_calls):
                call_id = tc.get("id", f"call_{step}_{i}")
                assistant_msg["tool_calls"].append({
                    "id": call_id,
                    "type": "function",
                    "function": {"name": tc["name"], "arguments": json.dumps(tc["arguments"])},
                })
            working_messages.append(assistant_msg)

            for i, tc in enumerate(tool_calls):
                call_id = tc.get("id", f"call_{step}_{i}")
                result_str = _run_tool_and_log(tc, step + 1, tool_log, run_id, user_id)
                working_messages.append({"role": "tool", "tool_call_id": call_id, "content": result_str})

        elif use_native and provider == "ollama":
            # Ollama: similar to OpenAI format
            assistant_msg = {"role": "assistant", "content": text or ""}
            assistant_msg["tool_calls"] = []
            for i, tc in enumerate(tool_calls):
                assistant_msg["tool_calls"].append({
                    "function": {"name": tc["name"], "arguments": tc["arguments"]},
                })
            working_messages.append(assistant_msg)

            for tc in tool_calls:
                result_str = _run_tool_and_log(tc, step + 1, tool_log, run_id, user_id)
                working_messages.append({"role": "tool", "content": result_str})

        else:
            # Text fallback: append results as user message
            if text:
                working_messages.append({"role": "assistant", "content": text})
            parts = []
            for tc in tool_calls:
                result_str = _run_tool_and_log(tc, step + 1, tool_log, run_id, user_id)
                parts.append(f'<tool_result name="{tc["name"]}">{result_str}</tool_result>')
            working_messages.append({"role": "user", "content": "\n".join(parts)})

    # Max steps reached
    last_text = text or ""
    if tool_log:
        last_text += "\n\n(Agent reached maximum step limit)"
    broadcast_event({"type": "_agent_done", "run_id": run_id, "user_id": user_id,
        "answer": last_text[:500], "total_steps": len(tool_log),
        "duration_ms": round((time.time() - loop_start) * 1000, 1), "timestamp": time.time()})
    return last_text, tool_log, total_usage


OPENAI_MODELS = [
    {"name": "gpt-5.2", "size": 0},
    {"name": "gpt-5.2-chat-latest", "size": 0},
    {"name": "gpt-5.2-pro", "size": 0},
    {"name": "gpt-5.2-codex", "size": 0},
    {"name": "gpt-4.1", "size": 0},
    {"name": "gpt-4.1-mini", "size": 0},
    {"name": "gpt-4.1-nano", "size": 0},
    {"name": "gpt-4o", "size": 0},
    {"name": "gpt-4o-mini", "size": 0},
    {"name": "o3", "size": 0},
    {"name": "o3-mini", "size": 0},
    {"name": "o4-mini", "size": 0},
]

ANTHROPIC_MODELS = [
    {"name": "claude-sonnet-4-20250514", "size": 0},
    {"name": "claude-haiku-4-5-20251001", "size": 0},
    {"name": "claude-3-5-sonnet-20241022", "size": 0},
    {"name": "claude-3-5-haiku-20241022", "size": 0},
    {"name": "claude-3-opus-20240229", "size": 0},
]


def list_models() -> list[dict]:
    with config_lock:
        provider = config.get("provider", "ollama")

    if provider == "openai":
        bearer = get_openai_bearer_token()
        if not bearer:
            return OPENAI_MODELS  # fallback list
        use_oauth = _using_oauth_token()
        if use_oauth:
            # OAuth: fetch models from ChatGPT backend
            models_url = f"{CHATGPT_BACKEND_URL}/models"
            headers = {"Authorization": f"Bearer {bearer}"}
            account_id = _get_chatgpt_account_id()
            if account_id:
                headers["ChatGPT-Account-ID"] = account_id
            try:
                req = Request(models_url, method="GET", headers=headers)
                resp = urlopen(req, timeout=5)
                data = json.loads(resp.read())
                models = data.get("data", [])
                chat_models = [
                    {"name": m["id"], "size": 0}
                    for m in models
                    if any(m["id"].startswith(p) for p in ("gpt-", "o1", "o3", "o4", "chatgpt-"))
                ]
                if chat_models:
                    return sorted(chat_models, key=lambda x: x["name"])
            except Exception as e:
                print(f"[openai] ChatGPT model list failed: {e}", file=sys.stderr)
            return CHATGPT_CODEX_MODELS
        else:
            with config_lock:
                base_url = config.get("openai_base_url", "https://api.openai.com/v1").rstrip("/")
            models_url = f"{base_url}/models"
        try:
            req = Request(
                models_url, method="GET",
                headers={"Authorization": f"Bearer {bearer}"},
            )
            resp = urlopen(req, timeout=5)
            data = json.loads(resp.read())
            models = data.get("data", [])
            # Filter to chat models
            chat_models = [
                {"name": m["id"], "size": 0}
                for m in models
                if any(m["id"].startswith(p) for p in ("gpt-", "o1", "o3", "o4", "chatgpt-"))
            ]
            if chat_models:
                return sorted(chat_models, key=lambda x: x["name"])
            return OPENAI_MODELS
        except Exception:
            return OPENAI_MODELS

    elif provider == "anthropic":
        return ANTHROPIC_MODELS

    elif provider == "openrouter":
        with config_lock:
            api_key = config["openrouter_api_key"]
        if not api_key:
            return OPENAI_MODELS + ANTHROPIC_MODELS  # show common models as fallback
        try:
            req = Request(
                "https://openrouter.ai/api/v1/models", method="GET",
                headers={"Authorization": f"Bearer {api_key}"},
            )
            resp = urlopen(req, timeout=5)
            data = json.loads(resp.read())
            models = data.get("data", [])
            return [{"name": m["id"], "size": 0} for m in models[:100]]  # cap at 100
        except Exception:
            return OPENAI_MODELS + ANTHROPIC_MODELS

    else:  # ollama
        with config_lock:
            ollama_url = config["ollama_url"]
        try:
            req = Request(f"{ollama_url}/api/tags", method="GET")
            resp = urlopen(req, timeout=5)
            data = json.loads(resp.read())
            models = data.get("models", [])
            return [{"name": m["name"], "size": m.get("size", 0)} for m in models]
        except Exception:
            return []


# ---------------------------------------------------------------------------
# Database connection module
# ---------------------------------------------------------------------------
_db_thread_local = threading.local()


def get_db_connection():
    """Return a DB connection or None, cached per-thread."""
    with config_lock:
        db_type = config.get("db_type", "")
        db_enabled = config.get("db_enabled", False)
    if not db_enabled or not db_type:
        return None

    # Check if we already have a connection for this thread
    existing = getattr(_db_thread_local, "conn", None)
    existing_type = getattr(_db_thread_local, "conn_type", "")
    if existing and existing_type == db_type:
        try:
            # Test if connection is still alive
            if db_type == "sqlite":
                existing.execute("SELECT 1")
            else:
                existing.execute("SELECT 1")
            return existing
        except Exception:
            try:
                existing.close()
            except Exception:
                pass
            _db_thread_local.conn = None

    try:
        if db_type == "mssql":
            if not HAS_PYODBC:
                return None
            with config_lock:
                server = config.get("db_mssql_server", "")
                database = config.get("db_mssql_database", "")
                user = config.get("db_mssql_user", "")
                password = config.get("db_mssql_password", "")
                driver = config.get("db_mssql_driver", "ODBC Driver 17 for SQL Server")
            conn_str = f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};UID={user};PWD={password}"
            conn = pyodbc.connect(conn_str, timeout=10)
            _db_thread_local.conn = conn
            _db_thread_local.conn_type = db_type
            return conn
        elif db_type == "sqlite":
            with config_lock:
                db_path = config.get("db_sqlite_path", "")
            if not db_path:
                return None
            # Resolve relative paths against data dir
            p = Path(db_path)
            if not p.is_absolute():
                p = DATA_DIR / db_path
            conn = sqlite3.connect(str(p), timeout=10)
            conn.row_factory = sqlite3.Row
            _db_thread_local.conn = conn
            _db_thread_local.conn_type = db_type
            return conn
    except Exception as e:
        print(f"[db] Connection error: {e}", file=sys.stderr)
    return None


def get_db_schema(force_refresh=False) -> str:
    """Discover tables + columns. Returns formatted string for LLM."""
    with config_lock:
        cached = config.get("db_schema_cache", "")
    if cached and not force_refresh:
        return cached

    conn = get_db_connection()
    if not conn:
        return ""

    with config_lock:
        db_type = config.get("db_type", "")

    schema_lines = []
    try:
        if db_type == "mssql":
            cursor = conn.cursor()
            cursor.execute(
                "SELECT TABLE_NAME, COLUMN_NAME, DATA_TYPE "
                "FROM INFORMATION_SCHEMA.COLUMNS ORDER BY TABLE_NAME, ORDINAL_POSITION"
            )
            current_table = None
            for row in cursor.fetchall():
                tbl, col, dtype = row[0], row[1], row[2]
                if tbl != current_table:
                    current_table = tbl
                    schema_lines.append(f"\nTable: {tbl}")
                schema_lines.append(f"  - {col} ({dtype})")
            cursor.close()
        elif db_type == "sqlite":
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'")
            tables = [r[0] if not isinstance(r, sqlite3.Row) else r["name"] for r in cursor.fetchall()]
            for tbl in tables:
                schema_lines.append(f"\nTable: {tbl}")
                cursor.execute(f"PRAGMA table_info(`{tbl}`)")
                for col_info in cursor.fetchall():
                    col_name = col_info[1] if not isinstance(col_info, sqlite3.Row) else col_info["name"]
                    col_type = col_info[2] if not isinstance(col_info, sqlite3.Row) else col_info["type"]
                    schema_lines.append(f"  - {col_name} ({col_type})")
            cursor.close()
    except Exception as e:
        print(f"[db] Schema error: {e}", file=sys.stderr)
        return f"Error reading schema: {e}"

    schema = "\n".join(schema_lines).strip()
    with config_lock:
        config["db_schema_cache"] = schema
    save_config()
    return schema


_WRITE_SQL_RE = re.compile(
    r"\b(INSERT|UPDATE|DELETE|DROP|ALTER|CREATE|TRUNCATE|REPLACE|MERGE|EXEC|EXECUTE|GRANT|REVOKE)\b",
    re.IGNORECASE,
)


def execute_db_query(sql: str, params=None) -> dict:
    """Execute a SELECT query safely. Returns result dict."""
    # Read-only check
    if _WRITE_SQL_RE.search(sql):
        return {"error": "Only SELECT queries are allowed", "columns": [], "rows": [], "row_count": 0, "duration_ms": 0}

    conn = get_db_connection()
    if not conn:
        return {"error": "No database connection", "columns": [], "rows": [], "row_count": 0, "duration_ms": 0}

    with config_lock:
        db_type = config.get("db_type", "")

    start = time.time()
    try:
        cursor = conn.cursor()
        if db_type == "mssql":
            cursor.execute(sql, params or ())
        else:
            cursor.execute(sql, params or ())

        if cursor.description is None:
            cursor.close()
            return {"error": "Query returned no results", "columns": [], "rows": [], "row_count": 0,
                    "duration_ms": round((time.time() - start) * 1000, 1)}

        columns = [desc[0] for desc in cursor.description]
        rows_raw = cursor.fetchmany(100)  # limit to 100 rows
        rows = [dict(zip(columns, row)) for row in rows_raw]
        row_count = len(rows)
        cursor.close()
        duration_ms = round((time.time() - start) * 1000, 1)

        # Log as event
        ingest_event({
            "type": "query",
            "sql": sql,
            "row_count": row_count,
            "duration_ms": duration_ms,
            "results": rows[:10],  # only store first 10 in event for feed
        })

        return {"columns": columns, "rows": rows, "row_count": row_count, "duration_ms": duration_ms}
    except Exception as e:
        duration_ms = round((time.time() - start) * 1000, 1)
        return {"error": str(e), "columns": [], "rows": [], "row_count": 0, "duration_ms": duration_ms}


# ---------------------------------------------------------------------------
# Embedding store
# ---------------------------------------------------------------------------
EMBEDDINGS_DB = DATA_DIR / "embeddings.db"


def init_embedding_db():
    """Create embedding tables if they don't exist."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(EMBEDDINGS_DB))
    conn.execute("""
        CREATE TABLE IF NOT EXISTS documents (
            id TEXT PRIMARY KEY,
            filename TEXT,
            content TEXT,
            chunk_count INTEGER,
            uploaded_by TEXT,
            uploaded_at REAL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS chunks (
            id TEXT PRIMARY KEY,
            doc_id TEXT REFERENCES documents(id),
            content TEXT,
            chunk_index INTEGER,
            embedding BLOB,
            token_count INTEGER
        )
    """)
    conn.commit()
    conn.close()


def _get_embedding_conn():
    """Get a connection to the embeddings database."""
    return sqlite3.connect(str(EMBEDDINGS_DB))


def embed_text(text: str) -> list:
    """Get embedding vector for text. Returns list of floats."""
    with config_lock:
        provider = config.get("embedding_provider", "auto")
        model = config.get("embedding_model", "")
        openai_key = config.get("openai_api_key", "")
        ollama_url = config.get("ollama_url", "http://localhost:11434")

    # Auto-detect provider
    if provider == "auto":
        if openai_key:
            provider = "openai"
        else:
            provider = "ollama"

    if provider == "openai":
        if not model:
            model = "text-embedding-3-small"
        bearer = get_openai_bearer_token()
        if not bearer:
            raise ValueError("No OpenAI API key available for embeddings")
        with config_lock:
            base_url = config.get("openai_base_url", "https://api.openai.com/v1").rstrip("/")
        body = json.dumps({"model": model, "input": text}).encode("utf-8")
        req = Request(
            f"{base_url}/embeddings", data=body,
            headers={"Content-Type": "application/json", "Authorization": f"Bearer {bearer}"},
            method="POST",
        )
        resp = urlopen(req, timeout=30)
        result = json.loads(resp.read())
        return result["data"][0]["embedding"]

    elif provider == "ollama":
        if not model:
            model = "nomic-embed-text"
        body = json.dumps({"model": model, "prompt": text}).encode("utf-8")
        req = Request(
            f"{ollama_url}/api/embeddings", data=body,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        resp = urlopen(req, timeout=60)
        result = json.loads(resp.read())
        return result.get("embedding", [])

    raise ValueError(f"Unknown embedding provider: {provider}")


def parse_excel(file_bytes: bytes) -> str:
    """Parse an Excel (.xlsx) file and return its content as text.

    Each sheet becomes a section with a header row repeated for context.
    Format: 'Header1 | Header2 | ...' on the first row, then data rows.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is not installed. Run: pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                all_rows.append(cells)
        if all_rows:
            header = all_rows[0]
            header_line = " | ".join(header)
            lines = [f"## {sheet_name}", f"Spalten: {header_line}"]
            for row_cells in all_rows[1:]:
                # Format as "Header1: Value1, Header2: Value2, ..."
                pairs = []
                for h, v in zip(header, row_cells):
                    if v.strip():
                        pairs.append(f"{h}: {v}")
                if pairs:
                    lines.append(", ".join(pairs))
            parts.append("\n".join(lines))
    wb.close()
    return "\n\n".join(parts)


def chunk_tabular_text(text: str, chunk_size: int = 500) -> list:
    """Chunk tabular text (from Excel) keeping section headers with each chunk."""
    sections = re.split(r"(?=^## )", text, flags=re.MULTILINE)
    chunks = []

    for section in sections:
        section = section.strip()
        if not section:
            continue
        lines = section.split("\n")
        # First two lines are "## SheetName" and "Spalten: ..." — keep as header
        header_lines = []
        data_lines = []
        for i, line in enumerate(lines):
            if i < 2 or line.startswith("## ") or line.startswith("Spalten:"):
                header_lines.append(line)
            else:
                data_lines.append(line)
        header = "\n".join(header_lines)

        current_chunk_lines = []
        current_tokens = _count_tokens(header)
        for dline in data_lines:
            line_tokens = _count_tokens(dline)
            if current_tokens + line_tokens > chunk_size and current_chunk_lines:
                chunks.append(header + "\n" + "\n".join(current_chunk_lines))
                current_chunk_lines = []
                current_tokens = _count_tokens(header)
            current_chunk_lines.append(dline)
            current_tokens += line_tokens
        if current_chunk_lines:
            chunks.append(header + "\n" + "\n".join(current_chunk_lines))

    return chunks if chunks else [text]


def chunk_text(text: str, chunk_size: int = 500) -> list:
    """Split text into chunks by paragraphs, respecting token limits."""
    paragraphs = re.split(r"\n\s*\n", text)
    chunks = []
    current_chunk = ""

    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        para_tokens = _count_tokens(para)
        if para_tokens > chunk_size:
            # Split long paragraph by sentences
            if current_chunk:
                chunks.append(current_chunk.strip())
                current_chunk = ""
            sentences = re.split(r"(?<=[.!?])\s+", para)
            for sent in sentences:
                if _count_tokens(current_chunk + " " + sent) > chunk_size and current_chunk:
                    chunks.append(current_chunk.strip())
                    current_chunk = sent
                else:
                    current_chunk = (current_chunk + " " + sent).strip()
            if current_chunk:
                chunks.append(current_chunk.strip())
                current_chunk = ""
        elif _count_tokens(current_chunk + "\n\n" + para) > chunk_size and current_chunk:
            chunks.append(current_chunk.strip())
            current_chunk = para
        else:
            current_chunk = (current_chunk + "\n\n" + para).strip()

    if current_chunk.strip():
        chunks.append(current_chunk.strip())

    return chunks


def store_document(filename: str, content: str, username: str) -> dict:
    """Chunk, embed, and store a document. Returns doc info."""
    with config_lock:
        cs = config.get("rag_chunk_size", 500)

    doc_id = str(uuid.uuid4())
    # Use tabular chunker for Excel files (keeps headers with every chunk)
    is_tabular = filename.lower().endswith((".xlsx", ".xls")) or content.lstrip().startswith("## ") and "Spalten:" in content[:500]
    if is_tabular:
        chunks = chunk_tabular_text(content, cs)
    else:
        chunks = chunk_text(content, cs)

    conn = _get_embedding_conn()
    try:
        conn.execute(
            "INSERT INTO documents (id, filename, content, chunk_count, uploaded_by, uploaded_at) VALUES (?,?,?,?,?,?)",
            (doc_id, filename, content, len(chunks), username, time.time()),
        )

        for i, chunk_content in enumerate(chunks):
            chunk_id = str(uuid.uuid4())
            try:
                embedding = embed_text(chunk_content)
                if HAS_NUMPY:
                    emb_blob = np.array(embedding, dtype=np.float32).tobytes()
                else:
                    emb_blob = struct.pack(f"{len(embedding)}f", *embedding)
            except Exception as e:
                print(f"[embed] Error embedding chunk {i}: {e}", file=sys.stderr)
                emb_blob = b""

            token_count = _count_tokens(chunk_content)

            conn.execute(
                "INSERT INTO chunks (id, doc_id, content, chunk_index, embedding, token_count) VALUES (?,?,?,?,?,?)",
                (chunk_id, doc_id, chunk_content, i, emb_blob, token_count),
            )

        conn.commit()
    finally:
        conn.close()

    return {"doc_id": doc_id, "filename": filename, "chunk_count": len(chunks)}


def search_embeddings(query: str, top_k: int = 5) -> list:
    """Embed query and find top-K similar chunks using cosine similarity."""
    if not HAS_NUMPY:
        return []

    try:
        query_emb = np.array(embed_text(query), dtype=np.float32)
    except Exception as e:
        print(f"[rag] Error embedding query: {e}", file=sys.stderr)
        return []

    conn = _get_embedding_conn()
    try:
        cursor = conn.execute("SELECT id, doc_id, content, embedding, token_count FROM chunks WHERE embedding != ''")
        results = []
        for row in cursor.fetchall():
            chunk_id, doc_id, content, emb_blob, token_count = row
            if not emb_blob:
                continue
            try:
                emb = np.frombuffer(emb_blob, dtype=np.float32)
                if len(emb) != len(query_emb):
                    continue
                # Cosine similarity
                dot = np.dot(query_emb, emb)
                norm = np.linalg.norm(query_emb) * np.linalg.norm(emb)
                if norm == 0:
                    continue
                similarity = float(dot / norm)
                results.append({
                    "chunk_id": chunk_id,
                    "doc_id": doc_id,
                    "content": content,
                    "similarity": round(similarity, 4),
                    "token_count": token_count,
                })
            except Exception:
                continue

        results.sort(key=lambda x: -x["similarity"])
        return results[:top_k]
    finally:
        conn.close()


def delete_document(doc_id: str) -> bool:
    """Remove document and all its chunks."""
    conn = _get_embedding_conn()
    try:
        conn.execute("DELETE FROM chunks WHERE doc_id = ?", (doc_id,))
        conn.execute("DELETE FROM documents WHERE id = ?", (doc_id,))
        conn.commit()
        return True
    except Exception as e:
        print(f"[embed] Error deleting doc {doc_id}: {e}", file=sys.stderr)
        return False
    finally:
        conn.close()


def get_embedding_stats() -> dict:
    """Return stats about the embedding store."""
    conn = _get_embedding_conn()
    try:
        doc_count = conn.execute("SELECT COUNT(*) FROM documents").fetchone()[0]
        chunk_count = conn.execute("SELECT COUNT(*) FROM chunks").fetchone()[0]
        db_size = EMBEDDINGS_DB.stat().st_size if EMBEDDINGS_DB.exists() else 0
        return {"documents": doc_count, "chunks": chunk_count, "db_size_bytes": db_size}
    except Exception:
        return {"documents": 0, "chunks": 0, "db_size_bytes": 0}
    finally:
        conn.close()


def list_documents() -> list:
    """List all documents with metadata."""
    conn = _get_embedding_conn()
    try:
        cursor = conn.execute("SELECT id, filename, chunk_count, uploaded_by, uploaded_at FROM documents ORDER BY uploaded_at DESC")
        return [
            {"id": r[0], "filename": r[1], "chunk_count": r[2], "uploaded_by": r[3], "uploaded_at": r[4]}
            for r in cursor.fetchall()
        ]
    except Exception:
        return []
    finally:
        conn.close()


def get_document_detail(doc_id: str) -> dict | None:
    """Get full document with content and all chunks."""
    conn = _get_embedding_conn()
    try:
        row = conn.execute("SELECT id, filename, content, chunk_count, uploaded_by, uploaded_at FROM documents WHERE id=?", (doc_id,)).fetchone()
        if not row:
            return None
        doc = {"id": row[0], "filename": row[1], "content": row[2], "chunk_count": row[3], "uploaded_by": row[4], "uploaded_at": row[5]}
        chunks_cursor = conn.execute("SELECT id, content, chunk_index, token_count FROM chunks WHERE doc_id=? ORDER BY chunk_index", (doc_id,))
        doc["chunks"] = [
            {"id": r[0], "content": r[1], "chunk_index": r[2], "token_count": r[3]}
            for r in chunks_cursor.fetchall()
        ]
        return doc
    except Exception:
        return None
    finally:
        conn.close()


def update_document(doc_id: str, filename: str | None, content: str | None, username: str) -> dict:
    """Update a document. If content changed, re-chunk and re-embed."""
    conn = _get_embedding_conn()
    try:
        row = conn.execute("SELECT id, filename, content, chunk_count FROM documents WHERE id=?", (doc_id,)).fetchone()
        if not row:
            return {"error": "Document not found"}

        old_filename, old_content, old_chunk_count = row[1], row[2], row[3]
        new_filename = filename if filename is not None else old_filename
        new_content = content if content is not None else old_content

        content_changed = (new_content != old_content)

        if content_changed:
            # Re-chunk and re-embed
            with config_lock:
                cs = config.get("rag_chunk_size", 500)
            chunks = chunk_text(new_content, cs)

            # Delete old chunks
            conn.execute("DELETE FROM chunks WHERE doc_id=?", (doc_id,))

            # Insert new chunks
            for i, chunk_content in enumerate(chunks):
                chunk_id = str(uuid.uuid4())
                try:
                    embedding = embed_text(chunk_content)
                    if HAS_NUMPY:
                        emb_blob = np.array(embedding, dtype=np.float32).tobytes()
                    else:
                        emb_blob = struct.pack(f"{len(embedding)}f", *embedding)
                except Exception as e:
                    print(f"[embed] Error embedding chunk {i}: {e}", file=sys.stderr)
                    emb_blob = b""

                token_count = len(chunk_content) // 4
                if HAS_TIKTOKEN:
                    try:
                        enc = tiktoken.get_encoding("cl100k_base")
                        token_count = len(enc.encode(chunk_content))
                    except Exception:
                        pass

                conn.execute(
                    "INSERT INTO chunks (id, doc_id, content, chunk_index, embedding, token_count) VALUES (?,?,?,?,?,?)",
                    (chunk_id, doc_id, chunk_content, i, emb_blob, token_count),
                )

            conn.execute("UPDATE documents SET filename=?, content=?, chunk_count=?, uploaded_by=?, uploaded_at=? WHERE id=?",
                         (new_filename, new_content, len(chunks), username, time.time(), doc_id))
        else:
            # Only rename
            conn.execute("UPDATE documents SET filename=? WHERE id=?", (new_filename, doc_id))

        conn.commit()
        return {"ok": True, "doc_id": doc_id, "filename": new_filename,
                "chunk_count": len(chunks) if content_changed else old_chunk_count,
                "re_embedded": content_changed}
    except Exception as e:
        return {"error": str(e)}
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Stats
# ---------------------------------------------------------------------------
def compute_stats() -> dict:
    now = time.time()
    one_hour_ago = now - 3600
    twenty_four_ago = now - 86400

    # Single pass over events
    total = 0
    chat_count = 0
    query_count = 0
    error_count = 0
    recent_users = set()
    recent_msgs = 0
    chat_durations = []
    query_durations = []
    questions = []
    hourly = defaultdict(int)
    type_counts = defaultdict(int)

    with events_lock:
        total = len(events)
        for e in events:
            etype = e.get("type", "unknown")
            type_counts[etype] += 1
            ts = e.get("timestamp", 0)
            uid = e.get("user_id")

            if etype == "chat":
                chat_count += 1
                dur = e.get("duration_ms")
                if dur:
                    chat_durations.append(dur)
                q = e.get("question", "")
                if q:
                    questions.append(q)
                if ts > twenty_four_ago:
                    hour = datetime.fromtimestamp(ts).strftime("%H:00")
                    hourly[hour] += 1
            elif etype == "query":
                query_count += 1
                dur = e.get("duration_ms")
                if dur:
                    query_durations.append(dur)
            elif etype == "error":
                error_count += 1

            if ts > one_hour_ago:
                if uid:
                    recent_users.add(uid)
                if etype == "chat":
                    recent_msgs += 1

    avg_response = (sum(chat_durations) / len(chat_durations)) if chat_durations else 0
    avg_query = (sum(query_durations) / len(query_durations)) if query_durations else 0
    error_rate = (error_count / total * 100) if total > 0 else 0
    topics = extract_topics(" ".join(questions), top_n=10)
    hourly_sorted = dict(sorted(hourly.items()))

    with moderation_lock:
        flagged_count = sum(1 for m in moderation.values() if m.get("flagged"))
        reviewed_count = sum(1 for m in moderation.values() if m.get("reviewed"))
        unreviewed = flagged_count - reviewed_count

    with users_lock:
        blocked_count = sum(1 for u in users.values() if u.get("status") == "blocked")
        total_users = len(users)

    return {
        "total_events": total,
        "total_chats": chat_count,
        "total_queries": query_count,
        "total_errors": error_count,
        "active_users": len(recent_users),
        "total_users": total_users,
        "blocked_users": blocked_count,
        "messages_last_hour": recent_msgs,
        "avg_response_ms": round(avg_response, 1),
        "avg_query_ms": round(avg_query, 1),
        "error_rate": round(error_rate, 2),
        "topics": topics,
        "hourly": hourly_sorted,
        "flagged": flagged_count,
        "unreviewed": unreviewed,
        "reviewed": reviewed_count,
        "type_counts": dict(type_counts),
    }


# ---------------------------------------------------------------------------
# SSE
# ---------------------------------------------------------------------------
def broadcast_event(evt: dict):
    data = json.dumps(evt, ensure_ascii=False)
    with sse_lock:
        alive = []
        for client_entry in sse_clients:
            q = client_entry[0]
            try:
                q.put_nowait(data)
                alive.append(client_entry)
            except queue.Full:
                pass  # drop dead clients
        if len(alive) < len(sse_clients):
            sse_clients[:] = alive


# ---------------------------------------------------------------------------
# HTTP Handler
# ---------------------------------------------------------------------------
class MonitorHandler(SimpleHTTPRequestHandler):

    def log_message(self, fmt, *args):
        msg = fmt % args
        if "GET /api/stream" in msg:
            return
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", file=sys.stderr)

    def _send_json(self, data, status=200, headers=None):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Access-Control-Allow-Origin", "*")
        if headers:
            for k, v in headers:
                self.send_header(k, v)
        self.end_headers()
        self.wfile.write(body)

    def _read_body(self) -> bytes:
        length = int(self.headers.get("Content-Length", 0))
        return self.rfile.read(length)

    def _cors_headers(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    # --- Auth helpers ---

    def _get_session(self) -> tuple[dict | None, str]:
        """Parse Cookie header, return (session_dict_or_None, session_id)."""
        cookie = self.headers.get("Cookie", "")
        sid = ""
        for part in cookie.split(";"):
            part = part.strip()
            if part.startswith("session="):
                sid = part[8:]
                break
        return get_session_by_id(sid), sid

    # Role hierarchy: admin > mitarbeiter > user
    _ROLE_LEVEL = {"admin": 3, "mitarbeiter": 2, "user": 1}

    def _require_auth(self, role: str = "user") -> dict | None:
        """Returns session if authorized, sends 401/403 and returns None otherwise.
        Role check uses hierarchy: admin > mitarbeiter > user."""
        session, _ = self._get_session()
        if not session:
            self._send_json({"error": "not_authenticated"}, 401)
            return None
        required_level = self._ROLE_LEVEL.get(role, 1)
        user_level = self._ROLE_LEVEL.get(session.get("role", "user"), 1)
        if user_level < required_level:
            self._send_json({"error": "forbidden"}, 403)
            return None
        return session

    def _session_cookie(self, sid: str) -> tuple[str, str]:
        return ("Set-Cookie", f"session={sid}; Path=/; HttpOnly; SameSite=Lax; Max-Age={SESSION_TTL}")

    def _clear_cookie(self) -> tuple[str, str]:
        return ("Set-Cookie", "session=; Path=/; HttpOnly; SameSite=Lax; Max-Age=0")

    def do_OPTIONS(self):
        self.send_response(204)
        self._cors_headers()
        self.end_headers()

    # ----- POST -----

    def do_POST(self):
        path = urlparse(self.path).path

        # Public routes (no auth required)
        if path == "/event":
            return self._handle_post_event()
        if path == "/api/auth/login":
            return self._handle_login()
        if path == "/api/auth/register":
            return self._handle_register()
        if path == "/api/auth/logout":
            return self._handle_logout()

        # User-level routes
        if path == "/api/chat":
            session = self._require_auth("user")
            if not session:
                return
            return self._handle_chat(session)
        if path == "/api/chat/upload":
            session = self._require_auth("user")
            if not session:
                return
            return self._handle_chat_file_upload(session)
        if path == "/api/chat/clear":
            session = self._require_auth("user")
            if not session:
                return
            return self._handle_chat_clear(session)

        # Admin-level routes
        if path == "/api/openai/auth/start":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_openai_auth_start()
        if path == "/api/openai/auth/callback":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_openai_auth_callback()
        if path == "/api/openrouter/exchange":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_openrouter_exchange()
        if path == "/api/clear":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_clear()
        if path == "/api/config":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_post_config()
        if path.startswith("/api/moderate/"):
            session = self._require_auth("admin")
            if not session:
                return
            event_id = path.split("/api/moderate/", 1)[1]
            return self._handle_moderate(event_id)
        if path.startswith("/api/accounts/"):
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_account_action(path)
        if path.startswith("/api/users/"):
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_user_action(path)

        # Database endpoints (admin-only)
        if path == "/api/db/test":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_db_test()
        if path == "/api/db/query":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_db_query()

        # Embedding endpoints (admin-only)
        if path == "/api/embeddings/upload":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_embeddings_upload(session)
        if path == "/api/embeddings/text":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_embeddings_text(session)

        self._send_json({"error": "Not found"}, 404)

    def do_DELETE(self):
        path = urlparse(self.path).path
        if path.startswith("/api/embeddings/documents/"):
            session = self._require_auth("admin")
            if not session:
                return
            doc_id = path.split("/api/embeddings/documents/", 1)[1].rstrip("/")
            return self._handle_embeddings_delete(doc_id)
        self._send_json({"error": "Not found"}, 404)

    def do_PUT(self):
        path = urlparse(self.path).path
        if path.startswith("/api/embeddings/documents/"):
            session = self._require_auth("admin")
            if not session:
                return
            doc_id = path.split("/api/embeddings/documents/", 1)[1].rstrip("/")
            return self._handle_embeddings_update(doc_id, session)
        self._send_json({"error": "Not found"}, 404)

    # --- Auth endpoint handlers ---

    def _handle_login(self):
        try:
            body = self._read_body()
            data = json.loads(body)
        except (json.JSONDecodeError, Exception) as e:
            self._send_json({"error": str(e)}, 400)
            return

        username = data.get("username", "").strip()
        password = data.get("password", "")
        if not username or not password:
            self._send_json({"error": "username and password required"}, 400)
            return

        with accounts_lock:
            acct = accounts.get(username)
            if not acct or not _verify_password(password, acct["password_hash"], acct["salt"]):
                self._send_json({"error": "invalid_credentials"}, 401)
                return
            # Auto-upgrade legacy SHA-256 hashes to PBKDF2 on successful login
            if _needs_rehash(acct["password_hash"]):
                acct["password_hash"] = _hash_password(password, acct["salt"])
                _rehash_save = True
            else:
                _rehash_save = False
            role = acct["role"]

        if _rehash_save:
            save_accounts()

        sid = create_session(username, role)
        self._send_json(
            {"ok": True, "user": {"username": username, "role": role}},
            headers=[self._session_cookie(sid)],
        )

    def _handle_register(self):
        try:
            body = self._read_body()
            data = json.loads(body)
        except (json.JSONDecodeError, Exception) as e:
            self._send_json({"error": str(e)}, 400)
            return

        username = data.get("username", "").strip()
        password = data.get("password", "")
        if not username or not password:
            self._send_json({"error": "username and password required"}, 400)
            return
        if len(username) < 2:
            self._send_json({"error": "username too short"}, 400)
            return
        if len(password) < 3:
            self._send_json({"error": "password too short"}, 400)
            return

        salt = secrets.token_hex(16)
        with accounts_lock:
            if username in accounts:
                self._send_json({"error": "username_taken"}, 409)
                return
            accounts[username] = {
                "username": username,
                "password_hash": _hash_password(password, salt),
                "salt": salt,
                "role": "user",
                "priority": "normal",
                "created_at": time.time(),
            }
        save_accounts()

        sid = create_session(username, "user")
        self._send_json(
            {"ok": True, "user": {"username": username, "role": "user"}},
            headers=[self._session_cookie(sid)],
        )

    def _handle_logout(self):
        session, sid = self._get_session()
        if sid:
            clear_chat_history(sid)
            delete_session(sid)
        self._send_json({"ok": True}, headers=[self._clear_cookie()])

    def _handle_openrouter_exchange(self):
        """Exchange an OpenRouter OAuth code for an API key."""
        try:
            body = self._read_body()
            data = json.loads(body)
        except (json.JSONDecodeError, Exception) as e:
            self._send_json({"error": str(e)}, 400)
            return

        code = data.get("code", "").strip()
        if not code:
            self._send_json({"error": "code is required"}, 400)
            return

        try:
            req_body = json.dumps({"code": code}).encode("utf-8")
            req = Request(
                "https://openrouter.ai/api/v1/auth/keys",
                data=req_body,
                headers={"Content-Type": "application/json"},
                method="POST",
            )
            resp = urlopen(req, timeout=10)
            result = json.loads(resp.read())
            api_key = result.get("key", "")
            if not api_key:
                self._send_json({"error": "No key returned from OpenRouter"}, 502)
                return

            with config_lock:
                config["openrouter_api_key"] = api_key
                config["provider"] = "openrouter"
            save_config()

            self._send_json({"ok": True, "provider": "openrouter"})
        except Exception as e:
            self._send_json({"error": f"OpenRouter exchange failed: {e}"}, 502)

    def _handle_chat_file_upload(self, session: dict):
        """POST /api/chat/upload — Upload a file and return parsed text for chat context."""
        ctype = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in ctype:
            self._send_json({"error": "Content-Type must be multipart/form-data"}, 400)
            return

        try:
            body = self._read_body()
            filename, file_bytes = _parse_multipart(ctype, body)

            if not file_bytes:
                self._send_json({"error": "No file content"}, 400)
                return
            if not filename:
                filename = "upload.txt"

            # Parse file content
            if filename.lower().endswith((".xlsx", ".xls")):
                if not HAS_OPENPYXL:
                    self._send_json({"error": "Excel not supported (openpyxl missing)"}, 400)
                    return
                text = parse_excel(file_bytes)
            else:
                text = file_bytes.decode("utf-8", errors="replace")

            # Truncate if too large (max ~8000 tokens ≈ 32000 chars)
            max_chars = 32000
            truncated = False
            if len(text) > max_chars:
                text = text[:max_chars]
                truncated = True

            self._send_json({
                "ok": True,
                "filename": filename,
                "content": text,
                "char_count": len(text),
                "truncated": truncated,
            })
        except Exception as e:
            self._send_json({"error": f"File parse failed: {e}"}, 500)

    def _handle_chat_clear(self, session: dict):
        """Clear chat history for the current session."""
        _, sid = self._get_session()
        if sid:
            clear_chat_history(sid)
        self._send_json({"ok": True})

    def _handle_openai_auth_start(self):
        """Generate PKCE verifier/challenge and return auth URL."""
        verifier = _pkce_code_verifier()
        challenge = _pkce_code_challenge(verifier)
        state = secrets.token_urlsafe(32)

        # Store pending PKCE state
        with _pkce_lock:
            # Clean old entries (> 10 min)
            now = time.time()
            expired = [k for k, v in _openai_pkce_pending.items()
                       if now - v["created_at"] > 600]
            for k in expired:
                del _openai_pkce_pending[k]
            _openai_pkce_pending[state] = {
                "verifier": verifier,
                "created_at": now,
            }

        # Must use port 1455 + /auth/callback — that's what's registered with the Codex client_id
        callback_url = f"http://localhost:{OPENAI_CALLBACK_PORT}/auth/callback"
        params = urlencode({
            "response_type": "code",
            "client_id": OPENAI_CLIENT_ID,
            "redirect_uri": callback_url,
            "scope": OPENAI_SCOPE,
            "code_challenge": challenge,
            "code_challenge_method": "S256",
            "state": state,
            "codex_cli_simplified_flow": "true",
            "id_token_add_organizations": "true",
        })
        auth_url = f"{OPENAI_AUTH_URL}?{params}"

        # Start temporary callback server on port 1455 to catch the redirect
        _start_oauth_callback_server(state)

        self._send_json({"auth_url": auth_url, "state": state})

    def _handle_openai_auth_callback(self):
        """Exchange authorization code + PKCE verifier for tokens, then obtain API key."""
        try:
            body = self._read_body()
            data = json.loads(body)
        except (json.JSONDecodeError, Exception) as e:
            self._send_json({"error": str(e)}, 400)
            return

        code = data.get("code", "").strip()
        state = data.get("state", "").strip()
        if not code or not state:
            self._send_json({"error": "code and state are required"}, 400)
            return

        # Look up PKCE verifier
        with _pkce_lock:
            pending = _openai_pkce_pending.pop(state, None)
        if not pending:
            self._send_json({"error": "invalid or expired state"}, 400)
            return

        verifier = pending["verifier"]
        callback_url = f"http://localhost:{OPENAI_CALLBACK_PORT}/auth/callback"

        try:
            # Step 1: Exchange auth code for tokens (access_token, id_token, refresh_token)
            token_body = urlencode({
                "grant_type": "authorization_code",
                "client_id": OPENAI_CLIENT_ID,
                "code": code,
                "redirect_uri": callback_url,
                "code_verifier": verifier,
            }).encode("utf-8")
            req = Request(
                OPENAI_TOKEN_URL, data=token_body,
                headers={"Content-Type": "application/x-www-form-urlencoded"},
                method="POST",
            )
            resp = urlopen(req, timeout=10)
            result = json.loads(resp.read())

            access_token = result.get("access_token", "")
            id_token = result.get("id_token", "")
            refresh_token = result.get("refresh_token", "")
            expires_in = result.get("expires_in", 3600)

            if not access_token:
                self._send_json({"error": "No access_token returned from OpenAI"}, 502)
                return

            # Step 2: Exchange id_token for an API key (like Codex CLI does)
            # This converts the ChatGPT subscription OAuth token into a usable API key
            api_key = ""
            if id_token:
                try:
                    exchange_body = urlencode({
                        "grant_type": "urn:ietf:params:oauth:grant-type:token-exchange",
                        "client_id": OPENAI_CLIENT_ID,
                        "requested_token": "openai-api-key",
                        "subject_token": id_token,
                        "subject_token_type": "urn:ietf:params:oauth:token-type:id_token",
                    }).encode("utf-8")
                    req2 = Request(
                        OPENAI_TOKEN_URL, data=exchange_body,
                        headers={"Content-Type": "application/x-www-form-urlencoded"},
                        method="POST",
                    )
                    resp2 = urlopen(req2, timeout=10)
                    result2 = json.loads(resp2.read())
                    api_key = result2.get("access_token", "")
                    print(f"[oauth] Token exchange succeeded, got API key: {bool(api_key)}", file=sys.stderr)
                except Exception as exc:
                    print(f"[oauth] Token exchange for API key failed: {exc}", file=sys.stderr)
                    if hasattr(exc, 'read'):
                        try:
                            err_body = exc.read().decode()
                            print(f"[oauth] Exchange error body: {err_body}", file=sys.stderr)
                        except Exception:
                            pass

            # Extract account ID from access_token JWT for ChatGPT backend API calls
            account_id = ""
            claims = _decode_jwt_payload(access_token)
            orgs = claims.get("organizations", [])
            if orgs and isinstance(orgs, list):
                if isinstance(orgs[0], dict):
                    account_id = orgs[0].get("id", "")
                elif isinstance(orgs[0], str):
                    account_id = orgs[0]
            if not account_id:
                account_id = claims.get("account_id", "") or claims.get("org_id", "")
            print(f"[oauth] JWT claims keys: {list(claims.keys())}", file=sys.stderr)
            # Log the auth claim for debugging account_id extraction
            auth_claim = claims.get("https://api.openai.com/auth", {})
            profile_claim = claims.get("https://api.openai.com/profile", {})
            print(f"[oauth] auth claim: {json.dumps(auth_claim) if isinstance(auth_claim, dict) else auth_claim}", file=sys.stderr)
            print(f"[oauth] profile claim: {json.dumps(profile_claim) if isinstance(profile_claim, dict) else profile_claim}", file=sys.stderr)
            # Try to get account_id from auth claim
            if not account_id and isinstance(auth_claim, dict):
                account_id = (auth_claim.get("chatgpt_account_id", "") or
                              auth_claim.get("account_id", "") or
                              auth_claim.get("organization_id", ""))
            if account_id:
                print(f"[oauth] Extracted account_id: {account_id[:8]}...", file=sys.stderr)
            else:
                print(f"[oauth] WARNING: No account_id found in JWT claims", file=sys.stderr)

            with config_lock:
                config["openai_oauth_token"] = access_token
                config["openai_refresh_token"] = refresh_token
                config["openai_token_expires"] = time.time() + expires_in - 60
                config["openai_chatgpt_account_id"] = account_id
                if api_key:
                    config["openai_api_key"] = api_key
                config["provider"] = "openai"
            save_config()

            self._send_json({"ok": True, "provider": "openai", "has_api_key": bool(api_key), "has_account_id": bool(account_id)})
        except Exception as e:
            self._send_json({"error": f"OpenAI token exchange failed: {e}"}, 502)

    # --- Event handlers ---

    def _handle_post_event(self):
        try:
            body = self._read_body()
            evt = json.loads(body)
        except (json.JSONDecodeError, Exception) as e:
            self._send_json({"error": str(e)}, 400)
            return

        # Check bot_enabled
        with config_lock:
            bot_on = config.get("bot_enabled", True)
        if not bot_on:
            self._send_json({"error": "bot_disabled"}, 503)
            return

        # Check if user is blocked
        uid = evt.get("user_id", "")
        if uid and not is_user_allowed(uid):
            self._send_json({"error": "user_blocked", "user_id": uid}, 403)
            return

        evt = ingest_event(evt)
        self._send_json({"ok": True, "id": evt["id"]})

    def _handle_chat(self, session: dict):
        # Check bot_enabled
        with config_lock:
            bot_on = config.get("bot_enabled", True)
        if not bot_on:
            self._send_json({"error": "bot_disabled"}, 503)
            return

        try:
            body = self._read_body()
            data = json.loads(body)
        except (json.JSONDecodeError, Exception) as e:
            self._send_json({"error": str(e)}, 400)
            return

        question = data.get("question", "").strip()
        file_context = data.get("file_context", "").strip()
        file_name = data.get("file_name", "").strip()
        if not question:
            self._send_json({"error": "question is required"}, 400)
            return

        # If file content is attached, prepend it to the question
        if file_context:
            file_label = f"[Datei: {file_name}]" if file_name else "[Hochgeladene Datei]"
            question = f"{file_label}\n{file_context}\n\n{question}"

        # Admin can override user_id; regular users use their session username
        if session["role"] == "admin" and data.get("user_id"):
            user_id = data["user_id"]
        else:
            user_id = session["username"]

        # Check if user is blocked
        if not is_user_allowed(user_id):
            self._send_json({"error": "user_blocked", "user_id": user_id}, 403)
            return

        with config_lock:
            model = data.get("model") or config["default_model"]
            system_prompt = data.get("system_prompt", config["system_prompt"])
            temperature = data.get("temperature", config["temperature"])
            db_enabled = config.get("db_enabled", False)
            rag_enabled = config.get("rag_enabled", False)
            rag_top_k = config.get("rag_top_k", 5)

        # DSGVO: Only mitarbeiter and admin get DB/RAG access
        user_role = session.get("role", "user")
        if user_role not in ("admin", "mitarbeiter"):
            db_enabled = False
            rag_enabled = False

        # Get session ID for chat history
        _, sid = self._get_session()

        # Append user message to session history
        append_chat_message(sid, "user", question, username=user_id, model=model)

        # Build system prompt with optional DB schema and RAG context
        system_parts = []
        # Default to German if no custom system prompt is set
        if system_prompt:
            system_parts.append(system_prompt)
        else:
            system_parts.append("Antworte immer auf Deutsch, es sei denn der Benutzer schreibt in einer anderen Sprache.")

        # RAG: inject relevant knowledge base chunks
        rag_chunks = []
        has_rag_context = False
        if rag_enabled:
            try:
                rag_chunks = search_embeddings(question, rag_top_k)
                relevant = [c for c in rag_chunks if c["similarity"] > 0.3]
                if relevant:
                    has_rag_context = True
                    ctx = "\n---\n".join(c["content"] for c in relevant)
                    system_parts.append(
                        "KNOWLEDGE BASE (use this to answer factual/company questions directly — do NOT use SQL for information found here):\n---\n" + ctx + "\n---"
                    )
            except Exception as e:
                print(f"[rag] Search error: {e}", file=sys.stderr)

        # DB: inject schema and instruct LLM to generate SQL
        if db_enabled:
            schema = get_db_schema()
            if schema:
                db_instruction = (
                    "You also have access to a database. Here is the schema:\n" + schema +
                    "\n\nIMPORTANT: Only generate a SQL query (wrapped in ```sql ... ```) when the user asks about "
                    "data, records, statistics, or information that would be stored in these database tables. "
                    "If the question can be answered from the knowledge base context above or from general knowledge, "
                    "answer directly WITHOUT SQL. After I execute a SQL query, I'll give you the results to summarize."
                )
                system_parts.append(db_instruction)

        # Build messages: combined system prompt + history
        messages = []
        if system_parts:
            messages.append({"role": "system", "content": "\n\n".join(system_parts)})
        messages.extend(get_chat_history(sid))

        total_start = time.time()
        total_usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
        tool_log = []

        # --- Agent mode ---
        with config_lock:
            agent_enabled = config.get("agent_enabled", False)
            agent_max_steps = config.get("agent_max_steps", 8)

        if agent_enabled:
            available_tools = _get_available_tools(user_role)
            if available_tools:
                # Add agent instruction to system prompt
                agent_sys = (
                    "You are an agent with access to tools. Use them when needed to answer questions accurately. "
                    "Always prefer using tools over guessing. If a tool call fails, try a different approach."
                )
                if messages and messages[0]["role"] == "system":
                    messages[0]["content"] += "\n\n" + agent_sys
                else:
                    messages.insert(0, {"role": "system", "content": agent_sys})

                try:
                    agent_run_id = str(uuid.uuid4())
                    answer, tool_log, agent_usage = run_agent_loop(
                        messages, model, temperature, available_tools,
                        max_steps=agent_max_steps, user_id=user_id, run_id=agent_run_id
                    )
                    for k in total_usage:
                        total_usage[k] += agent_usage.get(k, 0)
                except Exception as e:
                    broadcast_event({"type": "_agent_done", "run_id": agent_run_id, "user_id": user_id,
                        "answer": f"Error: {e}", "total_steps": 0, "error": True,
                        "duration_ms": round((time.time() - total_start) * 1000, 1), "timestamp": time.time()})
                    ingest_event({
                        "type": "error", "message": str(e),
                        "error_type": "agent_error", "user_id": user_id,
                    })
                    self._send_json({"error": str(e)}, 502)
                    return

                total_ms = round((time.time() - total_start) * 1000, 1)
                track_token_usage(user_id, total_usage)
                append_chat_message(sid, "assistant", answer, username=user_id, model=model)

                evt_data = {
                    "type": "chat", "user_id": user_id, "question": question,
                    "answer": answer, "duration_ms": total_ms, "model": model,
                }
                if total_usage.get("total_tokens"):
                    evt_data["token_usage"] = total_usage
                if tool_log:
                    evt_data["tool_log"] = tool_log
                evt = ingest_event(evt_data)

                resp = {
                    "ok": True, "id": evt["id"], "answer": answer,
                    "duration_ms": total_ms, "model": model,
                }
                if total_usage.get("total_tokens"):
                    resp["token_usage"] = total_usage
                if tool_log:
                    resp["tool_log"] = tool_log
                self._send_json(resp)
                return

        # --- Standard (non-agent) path ---
        try:
            answer, duration_ms, token_usage = call_llm(messages, model, temperature)
            for k in total_usage:
                total_usage[k] += token_usage.get(k, 0)
        except Exception as e:
            ingest_event({
                "type": "error", "message": str(e),
                "error_type": "ollama_error", "user_id": user_id,
            })
            self._send_json({"error": str(e)}, 502)
            return

        # NL-to-SQL: check if LLM returned a SQL block
        sql_executed = None
        if db_enabled and "```sql" in answer:
            sql_match = re.search(r"```sql\s*\n?(.*?)\n?\s*```", answer, re.DOTALL)
            if sql_match:
                sql_code = sql_match.group(1).strip()
                query_result = execute_db_query(sql_code)

                if query_result.get("error"):
                    # Error — ask LLM to explain
                    error_msg = f"SQL query failed: {query_result['error']}\nQuery: {sql_code}"
                    append_chat_message(sid, "assistant", answer, username=user_id, model=model)
                    append_chat_message(sid, "user", error_msg, username=user_id, model=model)
                    messages.append({"role": "assistant", "content": answer})
                    messages.append({"role": "user", "content": error_msg})
                    try:
                        answer, extra_ms, extra_usage = call_llm(messages, model, temperature)
                        duration_ms += extra_ms
                        for k in total_usage:
                            total_usage[k] += extra_usage.get(k, 0)
                    except Exception:
                        pass
                    sql_executed = {"sql": sql_code, "error": query_result["error"]}
                else:
                    # Success — ask LLM to summarize results
                    rows = query_result["rows"]
                    result_text = f"Query executed successfully. {query_result['row_count']} rows returned"
                    if rows:
                        # Format as markdown table for the LLM
                        cols = query_result["columns"]
                        header = " | ".join(cols)
                        sep = " | ".join("---" for _ in cols)
                        data_rows = "\n".join(" | ".join(str(r.get(c, "")) for c in cols) for r in rows[:50])
                        result_text += f":\n\n{header}\n{sep}\n{data_rows}"
                    result_text += "\n\nPlease summarize these results for the user."

                    append_chat_message(sid, "assistant", answer, username=user_id, model=model)
                    append_chat_message(sid, "user", result_text, username=user_id, model=model)
                    messages.append({"role": "assistant", "content": answer})
                    messages.append({"role": "user", "content": result_text})
                    try:
                        answer, extra_ms, extra_usage = call_llm(messages, model, temperature)
                        duration_ms += extra_ms
                        for k in total_usage:
                            total_usage[k] += extra_usage.get(k, 0)
                    except Exception:
                        pass
                    sql_executed = {"sql": sql_code, "row_count": query_result["row_count"],
                                    "duration_ms": query_result["duration_ms"]}

        total_ms = round((time.time() - total_start) * 1000, 1)

        # Track token usage
        track_token_usage(user_id, total_usage)

        # Append assistant response to session history
        append_chat_message(sid, "assistant", answer, username=user_id, model=model)

        evt_data = {
            "type": "chat", "user_id": user_id, "question": question,
            "answer": answer, "duration_ms": total_ms, "model": model,
        }
        if total_usage.get("total_tokens"):
            evt_data["token_usage"] = total_usage
        if sql_executed:
            evt_data["sql_executed"] = sql_executed
        evt = ingest_event(evt_data)

        resp = {
            "ok": True, "id": evt["id"], "answer": answer,
            "duration_ms": total_ms, "model": model,
        }
        if total_usage.get("total_tokens"):
            resp["token_usage"] = total_usage
        if sql_executed:
            resp["sql_executed"] = sql_executed
        self._send_json(resp)

    def _handle_clear(self):
        clear_all_events()
        broadcast_event({"type": "_clear"})
        self._send_json({"ok": True})

    def _handle_post_config(self):
        try:
            body = self._read_body()
            data = json.loads(body)
        except (json.JSONDecodeError, Exception) as e:
            self._send_json({"error": str(e)}, 400)
            return

        db_config_changed = False
        with config_lock:
            for key in ("ollama_url", "default_model", "system_prompt", "temperature",
                        "provider", "openai_base_url"):
                if key in data:
                    config[key] = data[key]
            # Only update API keys/tokens if a real value is sent (not masked); allow empty to clear
            for key in ("openai_api_key", "anthropic_api_key", "openrouter_api_key",
                         "openai_oauth_token", "openai_refresh_token", "db_mssql_password"):
                if key in data and not str(data[key]).startswith("..."):
                    config[key] = data[key]
            if "provider" in data and data["provider"] in ("ollama", "openai", "anthropic", "openrouter"):
                config["provider"] = data["provider"]
            if "bot_enabled" in data:
                config["bot_enabled"] = bool(data["bot_enabled"])
            # DB config keys
            for key in ("db_type", "db_mssql_server", "db_mssql_database", "db_mssql_user",
                        "db_mssql_driver", "db_sqlite_path"):
                if key in data:
                    if config.get(key) != data[key]:
                        db_config_changed = True
                    config[key] = data[key]
            if "db_enabled" in data:
                config["db_enabled"] = bool(data["db_enabled"])
            # RAG config keys
            if "rag_enabled" in data:
                config["rag_enabled"] = bool(data["rag_enabled"])
            for key in ("rag_top_k", "rag_chunk_size"):
                if key in data:
                    config[key] = int(data[key])
            for key in ("embedding_provider", "embedding_model"):
                if key in data:
                    config[key] = data[key]
            # Agent config keys
            if "agent_enabled" in data:
                config["agent_enabled"] = bool(data["agent_enabled"])
            if "agent_web_search" in data:
                config["agent_web_search"] = bool(data["agent_web_search"])
            if "agent_max_steps" in data:
                config["agent_max_steps"] = max(1, min(20, int(data["agent_max_steps"])))
        # Refresh schema cache if DB config changed
        if db_config_changed:
            with config_lock:
                config["db_schema_cache"] = ""
        save_config()
        with config_lock:
            self._send_json({"ok": True, "config": dict(config)})

    def _handle_moderate(self, event_id: str):
        try:
            body = self._read_body()
            data = json.loads(body)
        except (json.JSONDecodeError, Exception) as e:
            self._send_json({"error": str(e)}, 400)
            return

        with moderation_lock:
            mod = moderation.get(event_id, {})
            if "flagged" in data:
                mod["flagged"] = bool(data["flagged"])
            if "reviewed" in data:
                mod["reviewed"] = bool(data["reviewed"])
            if "note" in data:
                mod["note"] = str(data["note"])
            mod["updated_at"] = time.time()
            moderation[event_id] = mod

        persist_moderation(event_id, mod)
        broadcast_event({
            "type": "_moderation_update", "event_id": event_id,
            "moderation": mod, "timestamp": time.time(),
        })
        self._send_json({"ok": True, "moderation": mod})

    def _handle_user_action(self, path: str):
        """POST /api/users/<id>/block | /api/users/<id>/unblock | /api/users/<id>/update"""
        parts = path.split("/api/users/", 1)[1].split("/")
        if len(parts) < 2:
            self._send_json({"error": "Invalid path"}, 400)
            return

        user_id = parts[0]
        action = parts[1]

        if action == "block":
            with users_lock:
                if user_id in users:
                    users[user_id]["status"] = "blocked"
                else:
                    users[user_id] = {
                        "user_id": user_id, "status": "blocked",
                        "first_seen": time.time(), "last_seen": time.time(),
                        "message_count": 0, "error_count": 0, "note": "",
                        "priority": "normal",
                    }
            save_users()
            broadcast_event({"type": "_user_update", "user_id": user_id, "status": "blocked"})
            self._send_json({"ok": True, "user_id": user_id, "status": "blocked"})

        elif action == "unblock":
            with users_lock:
                if user_id in users:
                    users[user_id]["status"] = "active"
            save_users()
            broadcast_event({"type": "_user_update", "user_id": user_id, "status": "active"})
            self._send_json({"ok": True, "user_id": user_id, "status": "active"})

        elif action == "update":
            try:
                body = self._read_body()
                data = json.loads(body)
            except (json.JSONDecodeError, Exception) as e:
                self._send_json({"error": str(e)}, 400)
                return
            with users_lock:
                if user_id in users:
                    if "note" in data:
                        users[user_id]["note"] = str(data["note"])
                    if "status" in data and data["status"] in ("active", "blocked", "restricted"):
                        users[user_id]["status"] = data["status"]
                    if "priority" in data and data["priority"] in ("high", "normal", "low"):
                        users[user_id]["priority"] = data["priority"]
            save_users()
            with users_lock:
                self._send_json({"ok": True, "user": users.get(user_id, {})})

        else:
            self._send_json({"error": "Unknown action"}, 400)

    def _handle_account_action(self, path: str):
        """POST /api/accounts/<name>/role | priority | delete"""
        parts = path.split("/api/accounts/", 1)[1].split("/")
        if len(parts) < 2:
            self._send_json({"error": "Invalid path"}, 400)
            return

        username = parts[0]
        action = parts[1]

        if action == "role":
            try:
                body = self._read_body()
                data = json.loads(body)
            except (json.JSONDecodeError, Exception) as e:
                self._send_json({"error": str(e)}, 400)
                return
            new_role = data.get("role")
            if new_role not in ("admin", "mitarbeiter", "user"):
                self._send_json({"error": "role must be admin, mitarbeiter, or user"}, 400)
                return
            with accounts_lock:
                if username not in accounts:
                    self._send_json({"error": "account not found"}, 404)
                    return
                accounts[username]["role"] = new_role
            save_accounts()
            # Update any active sessions for this user
            with sessions_lock:
                for s in sessions.values():
                    if s["username"] == username:
                        s["role"] = new_role
            self._send_json({"ok": True, "username": username, "role": new_role})

        elif action == "priority":
            try:
                body = self._read_body()
                data = json.loads(body)
            except (json.JSONDecodeError, Exception) as e:
                self._send_json({"error": str(e)}, 400)
                return
            priority = data.get("priority")
            if priority not in ("high", "normal", "low"):
                self._send_json({"error": "priority must be high, normal, or low"}, 400)
                return
            with accounts_lock:
                if username not in accounts:
                    self._send_json({"error": "account not found"}, 404)
                    return
                accounts[username]["priority"] = priority
            save_accounts()
            self._send_json({"ok": True, "username": username, "priority": priority})

        elif action == "delete":
            session, _ = self._get_session()
            if session and session["username"] == username:
                self._send_json({"error": "cannot delete own account"}, 400)
                return
            with accounts_lock:
                if username not in accounts:
                    self._send_json({"error": "account not found"}, 404)
                    return
                # Cannot delete last admin
                if accounts[username]["role"] == "admin":
                    admin_count = sum(1 for a in accounts.values() if a["role"] == "admin")
                    if admin_count <= 1:
                        self._send_json({"error": "cannot delete last admin"}, 400)
                        return
                del accounts[username]
            save_accounts()
            # Remove sessions for deleted user
            with sessions_lock:
                to_del = [k for k, v in sessions.items() if v["username"] == username]
                for k in to_del:
                    del sessions[k]
            self._send_json({"ok": True, "username": username})

        else:
            self._send_json({"error": "Unknown action"}, 400)

    # ----- GET -----

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        params = parse_qs(parsed.query)

        # Public routes (no auth required)
        if path == "/auth/callback":
            return self._handle_openai_oauth_redirect(params)
        if path == "/api/auth/me":
            return self._handle_me()
        if path.startswith("/api/users/") and path.endswith("/check"):
            user_id = path.split("/api/users/", 1)[1].replace("/check", "")
            allowed = is_user_allowed(user_id)
            with users_lock:
                u = users.get(user_id, {})
                priority = u.get("priority", "normal")
            with config_lock:
                bot_enabled = config.get("bot_enabled", True)
            return self._send_json({
                "user_id": user_id, "allowed": allowed,
                "priority": priority, "bot_enabled": bot_enabled,
            })
        if path == "/api/health":
            return self._send_json({
                "status": "ok", "events": len(events),
                "users": len(users),
                "uptime_s": round(time.time() - SERVER_START),
            })

        # SSE stream (user-level auth)
        if path == "/api/stream":
            session = self._require_auth("user")
            if not session:
                return
            return self._handle_sse(session)

        # User-level endpoints
        if path == "/api/models":
            session = self._require_auth("user")
            if not session:
                return
            return self._handle_models()

        if path == "/api/usage":
            session = self._require_auth("user")
            if not session:
                return
            username = session["username"]
            usage = get_token_usage(username)
            # Only admins see all_users breakdown
            if session["role"] != "admin":
                usage.pop("all_users", None)
            return self._send_json(usage)

        # Admin-level endpoints
        if path == "/api/events":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_get_events(params)
        if path == "/api/conversations":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_conversations(params)
        if path == "/api/stats":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_stats()
        if path == "/api/moderation":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_get_moderation()
        if path == "/api/config":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_get_config()
        if path == "/api/users":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_get_users()
        if path == "/api/accounts":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_get_accounts()
        if path == "/api/admin/chats":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_admin_chats()
        if path.startswith("/api/admin/chats/"):
            session = self._require_auth("admin")
            if not session:
                return
            chat_sid = path.split("/api/admin/chats/", 1)[1].rstrip("/")
            return self._handle_admin_chat_detail(chat_sid)
        if path.startswith("/api/users/") and path.count("/") == 3:
            session = self._require_auth("admin")
            if not session:
                return
            user_id = path.split("/api/users/", 1)[1].rstrip("/")
            return self._handle_get_user(user_id)

        # Database endpoints (admin-only)
        if path == "/api/db/schema":
            session = self._require_auth("admin")
            if not session:
                return
            return self._handle_db_schema()

        # Embedding endpoints (read: mitarbeiter+, write: admin only)
        if path == "/api/embeddings/documents":
            session = self._require_auth("mitarbeiter")
            if not session:
                return
            return self._handle_embeddings_list()
        if path.startswith("/api/embeddings/documents/"):
            session = self._require_auth("mitarbeiter")
            if not session:
                return
            doc_id = path.split("/api/embeddings/documents/", 1)[1].rstrip("/")
            return self._handle_embeddings_detail(doc_id)
        if path == "/api/embeddings/search":
            session = self._require_auth("mitarbeiter")
            if not session:
                return
            return self._handle_embeddings_search(params)
        if path == "/api/embeddings/stats":
            session = self._require_auth("mitarbeiter")
            if not session:
                return
            return self._handle_embeddings_stats()

        # Static files (public)
        self._serve_static(path)

    def _handle_me(self):
        session, _ = self._get_session()
        if not session:
            self._send_json({"error": "not_authenticated"}, 401)
            return
        self._send_json({"user": {"username": session["username"], "role": session["role"]}})

    def _handle_openai_oauth_redirect(self, params: dict):
        """Serve a small HTML page that relays the OAuth code+state to the main dashboard."""
        code = (params.get("code") or [""])[0]
        state = (params.get("state") or [""])[0]
        error = (params.get("error") or [""])[0]
        # Redirect to dashboard root with the params so the JS can pick them up
        if error:
            html = f"""<!DOCTYPE html><html><body>
            <h2>OpenAI Login Failed</h2><p>Error: {error}</p>
            <p><a href="/">Back to Dashboard</a></p></body></html>"""
        else:
            # Redirect to / with code+state so the existing JS handler picks it up
            html = f"""<!DOCTYPE html><html><head>
            <meta http-equiv="refresh" content="0;url=/?code={code}&state={state}">
            </head><body>Redirecting...</body></html>"""
        body = html.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _serve_static(self, path: str):
        if path == "/" or path == "":
            path = "/index.html"
        filepath = (PUBLIC_DIR / path.lstrip("/")).resolve()
        # Prevent path traversal — resolved path must be inside PUBLIC_DIR
        if not str(filepath).startswith(str(PUBLIC_DIR.resolve())):
            self._send_json({"error": "Forbidden"}, 403)
            return
        if not filepath.exists() or not filepath.is_file():
            self._send_json({"error": "Not found"}, 404)
            return

        content = filepath.read_bytes()
        ct = "text/html"
        if path.endswith(".js"):
            ct = "application/javascript"
        elif path.endswith(".css"):
            ct = "text/css"
        elif path.endswith(".json"):
            ct = "application/json"
        elif path.endswith(".png"):
            ct = "image/png"
        elif path.endswith(".svg"):
            ct = "image/svg+xml"

        self.send_response(200)
        self.send_header("Content-Type", ct)
        self.send_header("Content-Length", str(len(content)))
        self.send_header("Cache-Control", "no-cache")
        self.end_headers()
        self.wfile.write(content)

    def _handle_sse(self, session: dict):
        self.send_response(200)
        self.send_header("Content-Type", "text/event-stream")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Connection", "keep-alive")
        self.end_headers()

        q: queue.Queue = queue.Queue(maxsize=500)
        session_info = {"username": session["username"], "role": session["role"]}
        client_entry = (q, session_info)
        with sse_lock:
            sse_clients.append(client_entry)

        try:
            # Send session info first
            session_data = json.dumps(session_info, ensure_ascii=False)
            self.wfile.write(f"event: session\ndata: {session_data}\n\n".encode())
            self.wfile.flush()

            if session["role"] == "admin":
                # Admin gets all events, moderation, and users
                with events_lock:
                    init = list(events)
                init_data = json.dumps(init, ensure_ascii=False)
                self.wfile.write(f"event: init\ndata: {init_data}\n\n".encode())
                self.wfile.flush()

                with moderation_lock:
                    mod_data = json.dumps(moderation, ensure_ascii=False)
                self.wfile.write(f"event: moderation\ndata: {mod_data}\n\n".encode())
                self.wfile.flush()

                with users_lock:
                    users_data = json.dumps(users, ensure_ascii=False)
                self.wfile.write(f"event: users\ndata: {users_data}\n\n".encode())
                self.wfile.flush()
            else:
                # Non-admin users get only their own events + system events
                username = session["username"]
                with events_lock:
                    user_events = [
                        e for e in events
                        if e.get("user_id") == username
                        or e.get("type") in ("system", "_clear")
                    ]
                init_data = json.dumps(user_events, ensure_ascii=False)
                self.wfile.write(f"event: init\ndata: {init_data}\n\n".encode())
                self.wfile.flush()

            while True:
                try:
                    data = q.get(timeout=15)
                    # Filter for non-admin users
                    if session["role"] != "admin":
                        try:
                            evt = json.loads(data)
                            evt_type = evt.get("type", "")
                            evt_uid = evt.get("user_id", "")
                            # Skip admin-only internal events
                            if evt_type.startswith("_") and evt_type != "_clear" and not evt_type.startswith("_agent_"):
                                continue
                            # Skip other users' events
                            if evt_uid and evt_uid != session["username"]:
                                continue
                        except json.JSONDecodeError:
                            pass
                    self.wfile.write(f"data: {data}\n\n".encode())
                    self.wfile.flush()
                except queue.Empty:
                    self.wfile.write(b": keepalive\n\n")
                    self.wfile.flush()
        except (BrokenPipeError, ConnectionResetError, OSError):
            pass
        finally:
            with sse_lock:
                try:
                    sse_clients.remove(client_entry)
                except ValueError:
                    pass

    def _handle_get_events(self, params: dict):
        type_filter = params.get("type", [None])[0]
        user_filter = params.get("user_id", [None])[0]
        flagged_only = params.get("flagged", [""])[0] == "true"
        limit = int(params.get("limit", [100])[0])
        offset = int(params.get("offset", [0])[0])

        with events_lock:
            filtered = list(events)

        if type_filter:
            filtered = [e for e in filtered if e.get("type") == type_filter]
        if user_filter:
            filtered = [e for e in filtered if e.get("user_id") == user_filter]
        if flagged_only:
            with moderation_lock:
                flagged_ids = {eid for eid, m in moderation.items() if m.get("flagged")}
            filtered = [e for e in filtered if e["id"] in flagged_ids]

        total = len(filtered)
        page = filtered[offset:offset + limit]

        self._send_json({"total": total, "events": page})

    def _handle_conversations(self, params: dict):
        with events_lock:
            chats = [e for e in events if e.get("type") == "chat"]

        convos: dict[str, dict] = {}
        for c in chats:
            uid = c.get("user_id", "anonymous")
            if uid not in convos:
                convos[uid] = {
                    "user_id": uid, "message_count": 0,
                    "first_message": c.get("timestamp", 0),
                    "last_message": c.get("timestamp", 0),
                    "messages": [],
                }
            convos[uid]["message_count"] += 1
            convos[uid]["last_message"] = max(convos[uid]["last_message"], c.get("timestamp", 0))
            convos[uid]["messages"].append(c)

        result = sorted(convos.values(), key=lambda x: -x["last_message"])
        self._send_json({"conversations": result})

    def _handle_stats(self):
        self._send_json(compute_stats())

    def _handle_get_moderation(self):
        with moderation_lock:
            self._send_json({"moderation": dict(moderation)})

    def _handle_models(self):
        models = list_models()
        self._send_json({"models": models})

    def _handle_get_config(self):
        with config_lock:
            cfg = dict(config)
        # Mask API keys and tokens — show only last 4 chars
        for key in ("openai_api_key", "anthropic_api_key", "openrouter_api_key",
                     "openai_oauth_token", "openai_refresh_token", "db_mssql_password"):
            val = cfg.get(key, "")
            if val:
                cfg[key] = "..." + val[-4:]
            else:
                cfg[key] = ""
        self._send_json({"config": cfg})

    def _handle_get_users(self):
        with users_lock:
            user_list = [dict(u) for u in users.values()]
        # Enrich with live event counts using pre-indexed lookups
        with events_lock:
            for u in user_list:
                uid = u["user_id"]
                user_evts = _events_by_user.get(uid, [])
                u["total_events"] = len(user_evts)
                u["total_chats"] = sum(1 for e in user_evts if e.get("type") == "chat")
                u["total_errors"] = sum(1 for e in user_evts if e.get("type") == "error")

        user_list.sort(key=lambda x: -x.get("last_seen", 0))
        self._send_json({"users": user_list})

    def _handle_get_user(self, user_id: str):
        self._send_json({"user": get_user_summary(user_id)})

    def _handle_admin_chats(self):
        """GET /api/admin/chats — list all active chat sessions with metadata."""
        with chat_sessions_lock:
            result = []
            for sid, messages in chat_sessions.items():
                meta = chat_session_meta.get(sid, {})
                result.append({
                    "session_id": sid,
                    "username": meta.get("username", ""),
                    "message_count": len(messages),
                    "created_at": meta.get("created_at", 0),
                    "last_activity": meta.get("last_activity", 0),
                    "model": meta.get("model", ""),
                })
        result.sort(key=lambda x: -x["last_activity"])
        self._send_json({"sessions": result})

    def _handle_admin_chat_detail(self, session_id: str):
        """GET /api/admin/chats/<session_id> — return full message history + metadata."""
        with chat_sessions_lock:
            messages = list(chat_sessions.get(session_id, []))
            meta = dict(chat_session_meta.get(session_id, {}))
        if not messages and not meta:
            self._send_json({"error": "session not found"}, 404)
            return
        self._send_json({"session_id": session_id, "meta": meta, "messages": messages})

    def _handle_get_accounts(self):
        with accounts_lock:
            acct_list = []
            for a in accounts.values():
                acct_list.append({
                    "username": a["username"],
                    "role": a["role"],
                    "priority": a.get("priority", "normal"),
                    "created_at": a.get("created_at"),
                })
        self._send_json({"accounts": acct_list})

    # --- Database endpoint handlers ---

    def _handle_db_test(self):
        """POST /api/db/test — Test database connection."""
        conn = get_db_connection()
        if conn is None:
            with config_lock:
                db_type = config.get("db_type", "")
                db_enabled = config.get("db_enabled", False)
            if not db_type:
                self._send_json({"ok": False, "error": "No database type configured"})
                return
            if not db_enabled:
                self._send_json({"ok": False, "error": "Database is not enabled"})
                return
            if db_type == "mssql" and not HAS_PYODBC:
                self._send_json({"ok": False, "error": "pyodbc is not installed. Run: pip install pyodbc"})
                return
            self._send_json({"ok": False, "error": "Could not connect to database — check your settings"})
            return

        schema = get_db_schema(force_refresh=True)
        table_count = schema.count("Table:")
        self._send_json({"ok": True, "tables": table_count, "schema_preview": schema[:500]})

    def _handle_db_schema(self):
        """GET /api/db/schema — Return cached schema."""
        schema = get_db_schema()
        self._send_json({"schema": schema})

    def _handle_db_query(self):
        """POST /api/db/query — Execute raw SQL (admin only, for testing)."""
        try:
            body = self._read_body()
            data = json.loads(body)
        except (json.JSONDecodeError, Exception) as e:
            self._send_json({"error": str(e)}, 400)
            return

        sql = data.get("sql", "").strip()
        if not sql:
            self._send_json({"error": "sql is required"}, 400)
            return

        result = execute_db_query(sql)
        self._send_json(result)

    # --- Embedding endpoint handlers ---

    def _handle_embeddings_list(self):
        """GET /api/embeddings/documents — List all documents."""
        docs = list_documents()
        self._send_json({"documents": docs})

    def _handle_embeddings_upload(self, session: dict):
        """POST /api/embeddings/upload — Upload a text file (multipart/form-data)."""
        ctype = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in ctype:
            self._send_json({"error": "Content-Type must be multipart/form-data"}, 400)
            return

        try:
            body = self._read_body()
            filename, file_bytes = _parse_multipart(ctype, body)

            if not file_bytes:
                self._send_json({"error": "No file content found"}, 400)
                return
            if not filename:
                filename = "upload.txt"

            # Handle Excel files
            if filename.lower().endswith((".xlsx", ".xls")):
                if not HAS_OPENPYXL:
                    self._send_json({"error": "Excel support requires openpyxl. Run: pip install openpyxl"}, 400)
                    return
                file_content = parse_excel(file_bytes)
            else:
                file_content = file_bytes.decode("utf-8", errors="replace")

            result = store_document(filename, file_content, session["username"])
            self._send_json({"ok": True, **result})
        except Exception as e:
            self._send_json({"error": f"Upload failed: {e}"}, 500)

    def _handle_embeddings_text(self, session: dict):
        """POST /api/embeddings/text — Upload raw text with title."""
        try:
            body = self._read_body()
            data = json.loads(body)
        except (json.JSONDecodeError, Exception) as e:
            self._send_json({"error": str(e)}, 400)
            return

        title = data.get("title", "").strip() or "Untitled"
        content = data.get("content", "").strip()
        if not content:
            self._send_json({"error": "content is required"}, 400)
            return

        try:
            result = store_document(title, content, session["username"])
            self._send_json({"ok": True, **result})
        except Exception as e:
            self._send_json({"error": f"Embedding failed: {e}"}, 500)

    def _handle_embeddings_delete(self, doc_id: str):
        """DELETE /api/embeddings/documents/<id> — Delete a document."""
        ok = delete_document(doc_id)
        if ok:
            self._send_json({"ok": True})
        else:
            self._send_json({"error": "Failed to delete document"}, 500)

    def _handle_embeddings_detail(self, doc_id: str):
        """GET /api/embeddings/documents/<id> — Full document with content and chunks."""
        doc = get_document_detail(doc_id)
        if not doc:
            self._send_json({"error": "Document not found"}, 404)
            return
        self._send_json({"document": doc})

    def _handle_embeddings_update(self, doc_id: str, session: dict):
        """PUT /api/embeddings/documents/<id> — Update document name/content, re-embed if changed."""
        try:
            body = self._read_body()
            data = json.loads(body)
        except (json.JSONDecodeError, Exception) as e:
            self._send_json({"error": str(e)}, 400)
            return

        filename = data.get("filename")
        content = data.get("content")
        if filename is None and content is None:
            self._send_json({"error": "Nothing to update"}, 400)
            return

        try:
            result = update_document(doc_id, filename, content, session["username"])
            if "error" in result:
                self._send_json(result, 400)
            else:
                self._send_json(result)
        except Exception as e:
            self._send_json({"error": f"Update failed: {e}"}, 500)

    def _handle_embeddings_search(self, params: dict):
        """GET /api/embeddings/search?q=... — Test search."""
        query = (params.get("q") or [""])[0]
        if not query:
            self._send_json({"error": "q parameter required"}, 400)
            return
        with config_lock:
            top_k = config.get("rag_top_k", 5)
        try:
            results = search_embeddings(query, top_k)
            self._send_json({"results": results})
        except Exception as e:
            self._send_json({"error": str(e)}, 500)

    def _handle_embeddings_stats(self):
        """GET /api/embeddings/stats — Total docs, chunks, DB size."""
        self._send_json(get_embedding_stats())


# ---------------------------------------------------------------------------
# Server
# ---------------------------------------------------------------------------
class ThreadedHTTPServer(ThreadingMixIn, HTTPServer):
    daemon_threads = True
    allow_reuse_address = True


SERVER_START = time.time()


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    load_config()
    load_users()
    load_events()
    load_accounts()
    ensure_default_admin()
    rebuild_users_from_events()
    save_users()
    init_embedding_db()
    print(f"[chatbot-monitor] Loaded {len(events)} events, {len(users)} users, {len(accounts)} accounts")
    print(f"[chatbot-monitor] Ollama: {config['ollama_url']}  Model: {config['default_model']}")
    print(f"[chatbot-monitor] DB: type={config.get('db_type','none')} enabled={config.get('db_enabled',False)}")
    print(f"[chatbot-monitor] RAG: enabled={config.get('rag_enabled',False)} embeddings_db={EMBEDDINGS_DB}")

    server = ThreadedHTTPServer(("0.0.0.0", PORT), MonitorHandler)
    print(f"[chatbot-monitor] Dashboard: http://localhost:{PORT}/")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[chatbot-monitor] Shutting down...")
        save_users()
        save_accounts()
        server.shutdown()


if __name__ == "__main__":
    main()
